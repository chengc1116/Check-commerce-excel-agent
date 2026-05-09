# -*- coding: utf-8 -*-
"""
Excel解析Agent — 项目书 → project_reviews 结构化JSON

核心理念：
  - Python只负责"搬运"：把Excel原始内容原样读出来
  - 全部交给LLM用语义理解去做字段映射
  - 一次性将4个sheet全部传入LLM
  - 输出纯JSON

用法:
    python scripts/excel_parsing_agent.py <xlsx文件路径>
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path

# ============================================================
# 环境设置
# ============================================================
sys.stdout.reconfigure(encoding="utf-8")

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT.parent))
os.chdir(PROJECT_ROOT.parent)

from dotenv import load_dotenv
# .env 在项目根目录（PROJECT_ROOT的父目录），而非product_review_agent目录
load_dotenv(PROJECT_ROOT.parent / ".env", override=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger("ExcelParsingAgent")


# ============================================================
# Excel原始内容提取（纯搬运，零理解）
# ============================================================

def read_excel_raw(file_path: str) -> list[dict]:
    """
    将Excel文件的所有sheet转为纯文本。
    返回 [{"sheet_name": str, "raw_text": str, "rows": int, "cols": int}, ...]
    不做任何字段理解，只忠实按行列输出，交给LLM处理。
    """
    from openpyxl import load_workbook
    
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    sheets_info = []

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("~$"):
            continue

        ws = wb[sheet_name]
        lines = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            row_values = []
            for cell in row:
                val = cell.value
                if val is not None:
                    row_values.append(str(val).strip())
                else:
                    row_values.append("")
            line = " | ".join(row_values)
            if line.strip(" |"):
                lines.append(line)
            else:
                lines.append("")

        raw_text = "\n".join(lines)

        sheets_info.append({
            "sheet_name": sheet_name,
            "raw_text": raw_text,
            "rows": ws.max_row,
            "cols": ws.max_column,
        })

        logger.info(f"  [{sheet_name}] {ws.max_row}行 x {ws.max_column}列, 文字约{len(raw_text)}字符")

    wb.close()
    return sheets_info


# ============================================================
# Prompt定义
# ============================================================

SYSTEM_PROMPT = """你是一个电商产品立项数据解析专家。你的任务是：将一份项目立项Excel表格的内容，转换成结构化JSON数据，用于存入「project_reviews」数据库表。

== 核心规则 ==
1. 你看到的Excel是原始表格内容，格式可能因人而异（列位置不同、写法不同、合并方式不同），你需要用语义理解来识别信息，不要依赖固定的列位置或列名。
2. 如果某个信息在多个地方出现，以最详细的那个为准。
3. 找不到的信息设为null，绝对不要编造数据。
4. 数字型字段尽量提取为数值（去掉单位符号）；日期保持原格式即可。
5. 返回纯JSON，不要markdown代码块。"""

# 通用模板（其他立项类型降级使用，待各类型专用模板补齐后可删除）
FALLBACK_PROMPT = """以下是Excel项目书的前4个sheet的完整内容：

{sheets_content}

---

请从以上Excel内容中提取以下字段，返回JSON。

== 字段提取说明（按大致位置提示，但请用语义理解识别，不要硬匹配列名）==
有可能给出的表格只有一个sheet，那么就只在sheet1中去匹配以下的信息。
【A. 基础信息】— 主要来自Sheet1
- project_name: 产品名称
- brand: 品牌
- product_code: 产品货号
- project_time: 立项时间
- design_time: 设计时间
- proofing_time: 打样时间
- launch_time: 上架时间
- categoryl1: 一级品类
- categoryl2: 二级品类
- categoryl3: 三级品类
- applicant: 负责人/申请人
- market_size: 市场规模描述
- estimated_sales: 目标销量额（提取数字，去掉单位）
【B. 市场与定价】— 主要来自Sheet2中 "价格/毛利" 的相关区域
- pricing: 定价（不变更形式）
- gfm: 毛利率（如72表示72%）
- ERP_price: ERP成本价（数字）
- core_config: 核心配置内容，不变更形式
- price_margin: 价格与毛利率（如"139.9单只/74%"，保留原格式）
- erp_cost: ERP成本（数字，如36.3）
【C. 设计要求】design_require 对象 — 来自Sheet1中"设计要求"相关区域:
  - content: 设计目的概述
  - outlook: 改外观/品牌的具体描述
  - material: 改材料的具体描述
  - function: 改功能的具体描述
【D. 对比产品信息】product_comparison 对象 — 来自Sheet2中"对手分析"区域:
  - comparison_name: 对手商品名称
  - selling_point: 对手的卖点（我方要复制的）
  - improving_point: 我方要超越/改进的点
【E. 使用人群】used_people 数组 — 来自Sheet3中"人群场景解析":
  针对表格内容自定义相关字段，根据表头去补充该部分字段
【F. 使用场景】used_scene 数组 — 来自Sheet3中"场景详细解析":
  针对表格内容自定义相关字段，根据表头去补充该部分字段
【G. 模块清单】module_list 数组 — 来自Sheet4（KANO+材质拆解）中的材质模块拆解:
  根据这一部分的表头自定义相关参数

== 注意事项 ==
- 找不到的字段设为null，数组字段找不到则设为空数组[]
- 返回纯JSON，不要加```json```包裹"""

# 提示词文件目录
PROMPTS_DIR = Path(__file__).resolve().parent / "prompts"


def load_task_prompt(task_type: str) -> str:
    """按立项类型加载对应的提示词md文件，找不到则返回通用模板"""
    prompt_file = PROMPTS_DIR / f"{task_type}.md"
    if prompt_file.exists():
        content = prompt_file.read_text(encoding="utf-8").strip()
        logger.info(f"[ExcelParsingAgent] 加载专用提示词: {prompt_file.name}")
        return content
    else:
        logger.warning(f"[ExcelParsingAgent] 未找到 {task_type}.md，使用通用模板")
        return FALLBACK_PROMPT



# ============================================================
# LLM 调用（含429重试）
# ============================================================

async def call_llm_with_retry(
    llm,
    system_prompt: str,
    user_prompt: str,
    max_retries: int = 3,
    base_delay: float = 10.0,
    use_fast_model: bool = False,
) -> dict:
    """调用LLM，自动处理429 Rate Limit重试"""
    call_fn = llm.acall_text_fast if use_fast_model else llm.acall_text
    for attempt in range(max_retries + 1):
        try:
            print(f"  [LLM调用] 第{attempt + 1}次尝试...")
            result = await call_fn(
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                response_format="json",
            )

            if isinstance(result, dict) and not result.get("_parse_error"):
                return result
            else:
                logger.warning(f"[ExcelParsingAgent] LLM返回解析异常: {type(result)}")
                return {"_error": f"parse_error: {str(result)[:200]}"}

        except Exception as e:
            error_name = type(e).__name__
            error_msg = str(e)

            if "429" in error_msg or "rate" in error_msg.lower() or "RateLimit" in error_name:
                if attempt < max_retries:
                    delay = base_delay * (2 ** attempt)
                    print(f"  [429限流] 等待{delay:.0f}秒后重试... ({attempt + 1}/{max_retries})")
                    await asyncio.sleep(delay)
                    continue
                else:
                    return {"_error": f"rate_limit_exceeded: {error_msg}"}

            elif "timeout" in error_msg.lower() or "Timeout" in error_name:
                if attempt < max_retries:
                    delay = base_delay
                    print(f"  [超时] 等待{delay:.0f}秒后重试... ({attempt + 1}/{max_retries})")
                    await asyncio.sleep(delay)
                    continue
                else:
                    return {"_error": f"timeout_exceeded: {error_msg}"}

            else:
                print(f"  [异常] {error_name}: {error_msg[:200]}")
                return {"_error": f"{error_name}: {error_msg}"}

    return {"_error": "max_retries_exceeded"}


# ============================================================
# 主解析流程
# ============================================================

async def parse_excel_to_project_review(file_path: str, task_type: str = "") -> dict:
    """Excel → 结构化JSON（一次性4个sheet传入LLM）

    Args:
        file_path: Excel文件路径
        task_type: 立项类型（如hot_upgrade），用于加载对应的提示词模板
    """
    start_time = time.time()
    file_path = str(file_path)
    filename = Path(file_path).name

    # Step 1: 读取Excel原始内容
    logger.info(f"[ExcelParsingAgent] 开始解析: {filename}, 类型: {task_type or '通用'}")
    sheets_data = read_excel_raw(file_path)
    logger.info(f"[ExcelParsingAgent] 共读取 {len(sheets_data)} 个sheet")

    # 只使用前4个sheet
    useful_sheets = sheets_data[:min(4, len(sheets_data))]
    logger.info(f"[ExcelParsingAgent] 使用前 {len(useful_sheets)} 个sheet")

    # Step 2: 组装4个sheet的内容
    sheets_parts = []
    for i, s in enumerate(useful_sheets):
        sheets_parts.append(
            f"== Sheet {i + 1}: {s['sheet_name']} ==\n\n{s['raw_text']}"
        )
    sheets_content = "\n\n\n".join(sheets_parts)

    total_chars = len(sheets_content)
    logger.info(f"[ExcelParsingAgent] 传入LLM内容总长: {total_chars}字符")
    print(f"\n  📊 传入LLM: {len(useful_sheets)}个sheet, 共{total_chars}字符")

    # Step 3: 按task_type加载提示词模板
    prompt_template = load_task_prompt(task_type) if task_type else FALLBACK_PROMPT
    user_prompt = prompt_template.format(sheets_content=sheets_content)

    # Step 4: 调用LLM
    from product_review_agent.agents.llm_client import get_llm_client
    llm = get_llm_client()

    if not llm.is_available:
        return {
            "_status": "llm_unavailable",
            "_error": "LLM未配置，请在.env中设置LLM_API_KEY",
            "filename": filename,
        }

    # Excel解析是简单的字段映射任务，使用快速模型即可
    original_max_tokens = llm.max_tokens
    llm.max_tokens = 8192

    try:
        print(f"\n  🚀 开始调用LLM（快速模型: {llm.fast_model}）...")
        logger.info(f"[ExcelParsingAgent] 调用快速模型: {llm.fast_model}")

        result = await call_llm_with_retry(
            llm,
            system_prompt=SYSTEM_PROMPT,
            user_prompt=user_prompt,
            use_fast_model=True,
        )
    finally:
        llm.max_tokens = original_max_tokens

    # Step 5: 适配层 — 将新字段映射回下游期望的格式
    if task_type == "category_gap":
        result = _adapt_category_gap(result)
    elif task_type in ("hot_upgrade", "competitor_upgrade", "low_sale_iterate"):
        result = _adapt_hot_upgrade(result)

    # Step 6: 补充元数据
    elapsed = round(time.time() - start_time, 1)
    has_error = result.get("_error")

    extracted = result if isinstance(result, dict) else {}
    extracted.update({
        "_status": "success" if not has_error else "failed",
        "filename": filename,
        "sheets_used": [s["sheet_name"] for s in useful_sheets],
        "_elapsed_seconds": elapsed,
        "_parsed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })

    if has_error:
        extracted["_error"] = has_error

    return extracted


def _adapt_hot_upgrade(data: dict) -> dict:
    """爆品升级适配层：仅做必要的字段转换，不重复搬运"""
    adapted = dict(data)

    # group_extra → group_extra_text（下游字段名不同）
    if data.get("group_extra"):
        adapted["group_extra_text"] = data["group_extra"]

    # pricing里提取gfm（毛利率）
    pricing = data.get("pricing", "")
    if pricing and not data.get("gfm"):
        import re
        m = re.search(r"[/／]\s*(\d+)\s*%?", pricing)
        if m:
            adapted["gfm"] = int(m.group(1))

    # erp_cost → ERP_price（下游字段名不同）
    if data.get("erp_cost") and not data.get("ERP_price"):
        adapted["ERP_price"] = data["erp_cost"]

    # 竞品信息 → product_comparison 对象（docx_generator读这个字段）
    if data.get("competitor_price") or data.get("competitor_url"):
        comparison = adapted.get("product_comparison", {})
        if data.get("competitor_price"):
            comparison["competitor_price"] = data["competitor_price"]
        if data.get("competitor_url"):
            comparison["comparison_url"] = data["competitor_url"]
        adapted["product_comparison"] = comparison

    return adapted


def _adapt_category_gap(data: dict) -> dict:
    """品类缺失适配层：将品类缺失专用字段映射为下游analyzer期望的格式"""
    adapted = dict(data)

    # ── 基础信息 other ──
    if data.get("base_other"):
        adapted["base_extra_text"] = data["base_other"]

    # ── 群体分析 → used_people / used_scene / target_audience / target_scenario ──
    people_analysis = data.get("people_analysis", "")
    scene_analysis = data.get("scene_analysis", "")
    group_other = data.get("group_other", "")

    if people_analysis:
        adapted["used_people"] = [{"raw_text": people_analysis}]
        adapted["target_audience"] = people_analysis
    if scene_analysis:
        adapted["used_scene"] = [{"raw_text": scene_analysis}]
        adapted["target_scenario"] = scene_analysis
    if group_other:
        adapted["group_extra_text"] = group_other

    # ── 竞品产品分析 → product_comparison 对象 ──
    comparison = {}
    if data.get("competitor_url"):
        comparison["comparison_url"] = data["competitor_url"]
    if data.get("competitor_price"):
        comparison["competitor_price"] = data["competitor_price"]
    if data.get("selling_point"):
        comparison["selling_point"] = data["selling_point"]
    if data.get("improving_point"):
        comparison["improving_point"] = data["improving_point"]
    if data.get("competitor_other"):
        comparison["competitor_other"] = data["competitor_other"]
    if comparison:
        adapted["product_comparison"] = comparison

    # 竞品名称/链接 → competitor_name（analyzer用）
    if data.get("competitor_url") and not data.get("competitor_name"):
        adapted["competitor_name"] = data.get("competitor_url")

    # ── 设计要求扁平字段 → design_require 对象 ──
    design_require = {}
    if data.get("design_purpose"):
        design_require["content"] = data["design_purpose"]
    if data.get("outlook"):
        design_require["outlook"] = data["outlook"]
    if data.get("material"):
        design_require["material"] = data["material"]
    if data.get("function"):
        design_require["function"] = data["function"]
    if data.get("design_other"):
        design_require["design_other"] = data["design_other"]
    if design_require:
        adapted["design_require"] = design_require

    # design_purpose → upgrade_direction（品类缺失没有升级方向，用设计目的替代）
    if data.get("design_purpose") and not data.get("upgrade_direction"):
        adapted["upgrade_direction"] = data["design_purpose"]

    # function → upgrade_function
    if data.get("function") and not data.get("upgrade_function"):
        adapted["upgrade_function"] = data["function"]

    # ── pricing里提取gfm（毛利率） ──
    pricing = data.get("pricing", "")
    if pricing and not data.get("gfm"):
        import re
        m = re.search(r"[/／]\s*(\d+)\s*%?", pricing)
        if m:
            adapted["gfm"] = int(m.group(1))

    # erp_cost → ERP_price
    if data.get("erp_cost") and not data.get("ERP_price"):
        adapted["ERP_price"] = data["erp_cost"]

    return adapted


# ============================================================
# CLI 入口
# ============================================================

async def main():
    args = sys.argv[1:]

    if not args or "--help" in args or "-h" in args:
        print("""
用法: python product_review_agent/parsers/excel_parsing_agent.py <xlsx文件路径> [--type 立项类型]

示例:
  python product_review_agent/parsers/excel_parsing_agent.py data/excel/xxx.xlsx --type hot_upgrade

立项类型:
  hot_upgrade       爆品升级
  category_gap      品类缺失
  competitor_upgrade 竞品升级
  low_sale_iterate   低销迭代
  (不指定则使用通用模板)

说明:
  一次性将前4个sheet传入LLM，用语义理解自动映射到数据库字段。
  输出纯JSON文件到 output/ 目录。
""")
        return

    file_path = args[0]
    task_type = ""
    if "--type" in args:
        idx = args.index("--type")
        if idx + 1 < len(args):
            task_type = args[idx + 1]

    print("\n" + "=" * 55)
    print("   Excel解析Agent — 项目书→结构化JSON")
    print("=" * 55 + "\n")

    result = await parse_excel_to_project_review(file_path, task_type=task_type)

    # 保存JSON
    output_dir = PROJECT_ROOT / "output"
    output_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = Path(file_path).stem

    json_path = output_dir / f"{stem}_parsed_{timestamp}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    # 打印结果
    print(json.dumps(result, ensure_ascii=False, indent=2))
    print(f"\n💾 JSON已保存: {json_path}")


if __name__ == "__main__":
    asyncio.run(main())
