# -*- coding: utf-8 -*-
"""
模板解析器 - 解析"产品研发要求输入表"Excel模板

支持两种数据来源:
1. 文本单元格 - 直接提取文字
2. 图片单元格 (DISPIMG公式/嵌入图片) - 提取图片供多模态LLM处理

设计思路:
- 扫描列B中的已知字段名,提取列C(及D)的值
- 列A中的文本作为板块标题(基本信息/九宫格/设计要求/具体情况)
- 合并单元格通过merge_map还原到左上角
- 只扫描前10列,避免异常列数(如seruna文件16360列)
"""

from __future__ import annotations

import logging
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, TwoCellAnchor

logger = logging.getLogger(__name__)


# ============================================================
# 字段映射: 列B中的关键词 → 输出JSON key
# ============================================================

FIELD_MAP: list[tuple[str, str]] = [
    # 立项信息
    ("立项时间", "立项时间"),
    ("设计时间", "设计时间"),
    ("打样时间", "打样时间"),
    ("上架时间", "上架时间"),
    # 基本信息
    ("一级品类", "category_l1"),
    ("二级品类", "category_l2"),
    ("三级品类", "category_l3"),
    ("产品名称", "product_name"),
    ("是否季节", "is_seasonal"),
    ("产品品牌", "brand"),
    ("负责人", "owner"),
    ("市场大小", "market_size"),
    ("对手销售额", "competitor_sales"),
    ("对手sku", "competitor_sku"),
    # 九宫格
    ("人群（落实", "target_audience"),
    ("人群", "target_audience"),
    ("场景（落实", "usage_scenarios"),
    ("场景", "usage_scenarios"),
    ("价格/毛利", "price_margin"),
    ("价格", "price_margin"),
    ("毛利", "price_margin"),
    ("对手是谁", "competitor_name"),
    # 设计要求
    ("设计目的概述", "design_purpose"),
    ("设计目的", "design_purpose"),
    ("改外观", "appearance_change"),
    ("改品牌", "appearance_change"),
    ("改材料", "material_change"),
    ("改功能", "function_change"),
    # 具体情况
    ("针对对手产品图片", "upgrade_details"),
    ("针对对手", "upgrade_details"),
    ("产品型号", "model_number"),
    ("ERP成本", "erp_cost"),
    ("ERP", "erp_cost"),
]

# "对手的卖点"需要特殊处理 - 后面跟(复制)或(超越)
COMPETITOR_PATTERNS: list[tuple[str, str]] = [
    ("对手的卖点（复制）", "competitor_strengths_copy"),
    ("对手的卖点（超越）", "competitor_advantage"),
    ("对手的卖点", "competitor_strengths_copy"),
]

# 板块标题关键词
SECTION_KEYWORDS: list[tuple[str, str]] = [
    ("基本信息", "基本信息"),
    ("9宫格", "九宫格目标"),
    ("九宫格", "九宫格目标"),
    ("设计要求", "设计要求"),
    ("具体情况", "具体情况"),
]

# 只扫描前N列(避免异常列数)
MAX_COL = 10


class TemplateParseResult:
    """模板解析结果"""

    def __init__(self, file_name: str):
        self.file_name = file_name
        self.sheet_name: str = ""
        self.data: dict[str, str] = {}  # field_key → text value
        self.image_cells: dict[str, dict] = {}  # field_key → image info
        self.warnings: list[str] = []
        self.sections: dict[str, list[str]] = {}  # section → [field_keys]

    def to_dict(self) -> dict:
        return {
            "file_name": self.file_name,
            "sheet_name": self.sheet_name,
            "data": self.data,
            "image_cells": list(self.image_cells.keys()),
            "warnings": self.warnings,
            "sections": self.sections,
        }


class TemplateParser:
    """产品研发要求输入表 模板解析器"""

    def parse(self, file_path: str | Path) -> TemplateParseResult:
        """
        解析Excel文件。
        
        自动寻找第一个有效工作表（跳过空表/~$开头），
        扫描已知字段名并提取对应值。
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        result = TemplateParseResult(file_path.name)

        wb = load_workbook(file_path, data_only=True)

        # 找到有效工作表
        valid_sheets = [s for s in wb.sheetnames if not s.startswith("~$")]
        if not valid_sheets:
            wb.close()
            raise ValueError(f"文件中没有有效工作表: {file_path}")

        for sheet_name in valid_sheets:
            ws = wb[sheet_name]
            result.sheet_name = sheet_name
            logger.info(
                f"解析工作表: {sheet_name} "
                f"(行={ws.max_row}, 列={ws.max_column})"
            )
            self._scan_sheet(ws, result)

        wb.close()
        return result

    def _scan_sheet(self, ws, result: TemplateParseResult):
        """扫描工作表,提取所有已知字段"""
        # 1. 构建合并单元格映射
        merge_map = self._build_merge_map(ws)

        # 2. 提取图片位置信息
        image_positions = self._get_image_positions(ws)

        # 3. 逐行扫描
        current_section = "未分类"
        seen_fields = set()  # 避免重复提取

        for row_idx, row in enumerate(ws.iter_rows(
            min_row=1, max_row=ws.max_row,
            min_col=1, max_col=min(MAX_COL, ws.max_column),
            values_only=False,
        ), start=1):
            cells = list(row)

            # 列A: 板块标题
            col_a = self._cell_text(cells, 0, merge_map)
            # 列B: 字段名
            col_b = self._cell_text(cells, 1, merge_map)
            # 列C: 字段值
            col_c_raw = self._cell_raw(cells, 2, merge_map)
            col_c = self._cell_text(cells, 2, merge_map)
            # 列D: 额外值
            col_d = self._cell_text(cells, 3, merge_map)

            # 跳过标题行
            if "产品研发要求输入表" in col_a or "产品研发要求输入表" in col_b:
                continue
            if "立项编号" in col_a:
                continue

            # 检查板块标题
            if col_b:  # 列B有内容时,列A是板块标题
                for kw, sec_name in SECTION_KEYWORDS:
                    if kw in col_a:
                        current_section = sec_name
                        break

            if not col_b:
                continue

            # 匹配字段名
            field_key = self._match_field(col_b)

            if not field_key or field_key in seen_fields:
                continue

            # 提取值
            value = col_c if col_c else ""
            extra = col_d if col_d else ""

            # 检查是否是图片单元格
            is_image = self._is_image_cell(col_c_raw, cells, 2, row_idx, image_positions)

            if is_image:
                # 尝试提取图片
                img_data = self._extract_image_at(ws, cells, 2, row_idx, image_positions)
                result.image_cells[field_key] = {
                    "row": row_idx,
                    "col": 3,  # 1-based
                    "formula": col_c_raw,
                    "image_data": img_data,  # bytes or None
                }
                result.warnings.append(
                    f"字段 [{col_b}] 包含图片,需要多模态解析"
                )
                value = "[图片 - 需多模态解析]"
            else:
                # 合并C和D列的值
                if value and extra and extra not in value:
                    value = f"{value}\n{extra}"
                elif not value and extra:
                    value = extra

            result.data[field_key] = value
            seen_fields.add(field_key)

            # 记录板块归属
            result.sections.setdefault(current_section, []).append(field_key)

            logger.debug(
                f"  [{current_section}] {col_b} -> {field_key} = {value[:60]}..."
            )

        # 4. 为没有匹配到板块的字段归类
        for key in result.data:
            in_any = any(key in v for v in result.sections.values())
            if not in_any:
                result.sections.setdefault("其他信息", []).append(key)

    def _build_merge_map(self, ws) -> dict:
        """构建合并单元格映射: (row, col) → text"""
        merge_map = {}
        for merged_range in ws.merged_cells.ranges:
            top_left = merged_range.start_cell
            text = str(ws.cell(top_left.row, top_left.column).value or "")
            if text.strip():
                for r in range(merged_range.min_row, merged_range.max_row + 1):
                    for c in range(merged_range.min_col, merged_range.max_col + 1):
                        merge_map[(r, c)] = text.strip()
        return merge_map

    def _cell_raw(self, cells, col_idx: int, merge_map: dict) -> str:
        """获取单元格原始值(含公式)"""
        if col_idx >= len(cells):
            return ""
        cell = cells[col_idx]
        # 优先返回value,其次返回合并单元格文本
        if cell.value is not None:
            return str(cell.value).strip()
        key = (cell.row, cell.column)
        return merge_map.get(key, "")

    def _cell_text(self, cells, col_idx: int, merge_map: dict) -> str:
        """获取单元格文本(过滤公式)"""
        raw = self._cell_raw(cells, col_idx, merge_map)
        if not raw:
            return ""
        # DISPIMG公式不算文本
        if "=DISPIMG" in raw.upper():
            return ""
        return raw.strip()

    def _match_field(self, col_b: str) -> Optional[str]:
        """匹配列B中的字段名,返回field_key"""
        # 先检查特殊的"对手的卖点"
        for pattern, key in COMPETITOR_PATTERNS:
            if pattern in col_b:
                return key

        # 常规字段匹配
        for pattern, key in FIELD_MAP:
            if pattern in col_b:
                return key

        return None

    def _is_image_cell(
        self, col_c_raw: str, cells, col_idx: int,
        row_idx: int, image_positions: set
    ) -> bool:
        """判断单元格是否包含图片"""
        # DISPIMG公式
        if "=DISPIMG" in col_c_raw.upper():
            return True
        # 该位置有嵌入图片
        if col_idx < len(cells):
            cell = cells[col_idx]
            if (cell.row, cell.column) in image_positions:
                return True
        return False

    def _get_image_positions(self, ws) -> set:
        """获取所有嵌入图片的单元格位置"""
        positions = set()
        try:
            for img in ws._images:
                anchor = img.anchor
                row, col = 0, 0
                if isinstance(anchor, OneCellAnchor):
                    row = anchor._from.row
                    col = anchor._from.col
                elif isinstance(anchor, TwoCellAnchor):
                    row = anchor._from.row
                    col = anchor._from.col
                positions.add((row + 1, col + 1))  # 转为1-based
                logger.debug(f"  发现图片: row={row+1}, col={col+1}, size={img.width}x{img.height}")
        except Exception as e:
            logger.warning(f"获取图片位置失败: {e}")
        return positions

    def _extract_image_at(
        self, ws, cells, col_idx: int,
        row_idx: int, image_positions: set
    ) -> Optional[bytes]:
        """提取指定位置的嵌入图片,返回bytes(失败返回None)"""
        if col_idx >= len(cells):
            return None
        cell = cells[col_idx]
        cell_pos = (cell.row, cell.column)

        try:
            for img in ws._images:
                anchor = img.anchor
                img_row, img_col = 0, 0
                if isinstance(anchor, OneCellAnchor):
                    img_row = anchor._from.row
                    img_col = anchor._from.col
                elif isinstance(anchor, TwoCellAnchor):
                    img_row = anchor._from.row
                    img_col = anchor._from.col

                # 检查图片是否覆盖此单元格
                if img_row + 1 == cell.row and img_col + 1 == cell.column:
                    buf = BytesIO()
                    try:
                        img_data = img._data()
                        if img_data:
                            return img_data
                    except AttributeError:
                        pass
                    # 备选: 通过ref保存
                    try:
                        img.save(buf)
                        return buf.getvalue()
                    except Exception:
                        pass
        except Exception as e:
            logger.warning(f"提取图片失败 (row={row_idx}): {e}")

        return None
