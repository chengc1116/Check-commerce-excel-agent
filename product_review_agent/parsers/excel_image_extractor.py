# -*- coding: utf-8 -*-
"""
Excel Sheet 图片提取器

支持两种图片类型：
1. 浮动图片 — 标准Excel drawing锚点（twoCellAnchor / oneCellAnchor）
2. 嵌入单元格图片 — WPS DISPIMG 公式 → cellimages.xml 链路

用法:
    from excel_image_extractor import extract_sheet_images

    images = extract_sheet_images("xxx.xlsx", sheet_name="4-竞品分析")
    # 返回: [{"cell": "B5", "type": "floating", "format": "png", "bytes": b"..."}, ...]

    # 保存到文件
    save_images(images, output_dir="output/images")
"""

from __future__ import annotations

import base64
import re
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Optional


# Excel XML 命名空间
NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "c": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "etc": "urn:schemas-microsoft-com:office:excel",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def _strip_ns(tag: str) -> str:
    """去掉XML命名空间前缀，如 {http://...}row → row"""
    return tag.split("}")[-1] if "}" in tag else tag


def _find_all_elements(root, local_name: str):
    """在XML树中查找所有指定local name的元素（忽略命名空间）"""
    results = []
    for elem in root.iter():
        if _strip_ns(elem.tag) == local_name:
            results.append(elem)
    return results


# ============================================================
# Sheet索引 → 文件路径映射
# ============================================================

def _get_sheet_file_map(zf: zipfile.ZipFile) -> dict:
    """
    构建 {sheet名: {sheet_xml, drawing_xml_list, rels_xml}} 的映射
    
    一个sheet可能关联多个drawing文件（通过rels）。
    """
    from openpyxl import load_workbook
    
    # 用openpyxl获取sheet名和索引的映射
    wb = load_workbook(zf.filename, read_only=True, data_only=False)
    sheet_names = wb.sheetnames
    wb.close()
    
    all_names_in_zip = zf.namelist()
    
    sheet_map = {}
    for idx, name in enumerate(sheet_names, 1):
        sheet_xml_path = f"xl/worksheets/sheet{idx}.xml"
        rels_path = f"xl/worksheets/_rels/sheet{idx}.xml.rels"
        
        drawings = []  # 可能多个drawing
        
        # 从rels文件找drawing关联
        if rels_path in all_names_in_zip:
            rels_content = zf.read(rels_path).decode("utf-8")
            for m in re.finditer(r'Target="([^"]*drawing[^"]*)"', rels_content, re.IGNORECASE):
                target = m.group(1)
                if target.startswith(".."):
                    drawing_xml = "xl/" + target.replace("../", "")
                elif target.startswith("xl/"):
                    drawing_xml = target
                else:
                    drawing_xml = f"xl/worksheets/{target}"
                
                # drawing的rels文件（注意用正斜杠，zip内部统一用/）
                drawing_name = drawing_xml.split("/")[-1]
                drawing_dir = "/".join(drawing_xml.split("/")[:-1])
                drawing_rels = f"{drawing_dir}/_rels/{drawing_name}.rels"
                
                drawings.append({
                    "drawing_xml": drawing_xml,
                    "drawing_rels": drawing_rels if drawing_rels in all_names_in_zip else None,
                })
        
        sheet_map[name] = {
            "index": idx,
            "sheet_xml": sheet_xml_path,
            "drawings": drawings,
        }
    
    return sheet_map


# ============================================================
# 类型1: 浮动图片提取（标准drawing锚点）
# ============================================================

def _extract_floating_images(zf: zipfile.ZipFile, sheet_info: dict) -> list[dict]:
    """
    从drawing XML中提取浮动图片。
    
    链路: drawing.xml → anchor → pic → blip r:embed → drawing.rels → media文件
    """
    images = []
    
    for drawing_info in sheet_info.get("drawings", []):
        drawing_xml_path = drawing_info.get("drawing_xml")
        if not drawing_xml_path or drawing_xml_path not in zf.namelist():
            continue
        
        # 解析drawing rels: rId → media文件路径
        rId_to_media = {}
        drawing_rels_path = drawing_info.get("drawing_rels")
        if drawing_rels_path and drawing_rels_path in zf.namelist():
            rels_content = zf.read(drawing_rels_path).decode("utf-8")
            for m in re.finditer(r'Id="(rId\d+)"[^>]*Target="([^"]*)"', rels_content):
                rId = m.group(1)
                target = m.group(2)
                if target.startswith(".."):
                    media_path = "xl/" + target.replace("../", "")
                elif target.startswith("xl/"):
                    media_path = target
                else:
                    media_path = f"xl/{target}"
                rId_to_media[rId] = media_path
        
        # 解析drawing XML
        drawing_content = zf.read(drawing_xml_path).decode("utf-8")
        root = ET.fromstring(drawing_content)
        
        # 处理 twoCellAnchor 和 oneCellAnchor
        for anchor in _find_all_elements(root, "twoCellAnchor") + _find_all_elements(root, "oneCellAnchor"):
            # 获取锚定位置（起始单元格）
            from_elem = _find_all_elements(anchor, "from")
            if not from_elem:
                continue
            
            col_elem = _find_all_elements(from_elem[0], "col")
            row_elem = _find_all_elements(from_elem[0], "row")
            if not col_elem or not row_elem:
                continue
            
            col = int(col_elem[0].text or "0")
            row = int(row_elem[0].text or "0")
            
            # 列号转字母 (0→A, 1→B, ...)
            col_letter = ""
            n = col
            while True:
                col_letter = chr(65 + n % 26) + col_letter
                n = n // 26 - 1
                if n < 0:
                    break
            
            cell_ref = f"{col_letter}{row + 1}"
            
            # 找 blip 的 r:embed
            blip_elems = _find_all_elements(anchor, "blip")
            if not blip_elems:
                continue
            
            rId = None
            for attr_name, attr_val in blip_elems[0].attrib.items():
                if attr_name.endswith("}embed") or attr_name == "embed":
                    rId = attr_val
                    break
            
            if not rId or rId not in rId_to_media:
                continue
            
            media_path = rId_to_media[rId]
            if media_path not in zf.namelist():
                continue
            
            # 读取图片二进制
            image_bytes = zf.read(media_path)
            fmt = Path(media_path).suffix.lstrip(".").lower()
            if fmt == "jpeg":
                fmt = "jpg"
            
            images.append({
                "cell": cell_ref,
                "type": "floating",
                "format": fmt,
                "bytes": image_bytes,
                "media_path": media_path,
            })
    
    return images


# ============================================================
# 类型2: 嵌入单元格图片（WPS DISPIMG → cellimages.xml）
# ============================================================

def _extract_cell_images(zf: zipfile.ZipFile, sheet_info: dict) -> list[dict]:
    """
    从WPS单元格图片中提取（DISPIMG公式 → cellimages.xml链路）。
    
    链路:
    sheet单元格 =DISPIMG("ID_xxx",1) 
        → cellimages.xml 找 name="ID_xxx" → blip r:embed
        → cellimages.xml.rels 找 rId → media文件
    """
    sheet_xml_path = sheet_info["sheet_xml"]
    if sheet_xml_path not in zf.namelist():
        return []
    
    # Step 1: 从sheet XML中找出所有DISPIMG单元格
    sheet_content = zf.read(sheet_xml_path).decode("utf-8")
    dispimg_cells = {}  # {dispimg_id: cell_ref}
    
    # WPS单元格图片公式格式:
    # <f>_xlfn.DISPIMG("ID_xxx",1)</f>  或
    # <f>_xlfn.DISPIMG(&quot;ID_xxx&quot;,1)</f>
    # 先把HTML实体替换
    clean_content = sheet_content.replace("&quot;", '"')
    
    for m in re.finditer(
        r'<c r="([A-Z]+\d+)"[^>]*>.*?<f>[^<]*DISPIMG\("([^"]+)"',
        clean_content, re.DOTALL
    ):
        cell_ref = m.group(1)
        dispimg_id = m.group(2)
        dispimg_cells[dispimg_id] = cell_ref
    
    if not dispimg_cells:
        return []
    
    # Step 2: 解析 cellimages.xml.rels → rId到media路径
    rId_to_media = {}
    cellimages_rels_path = "xl/_rels/cellimages.xml.rels"
    if cellimages_rels_path in zf.namelist():
        rels_content = zf.read(cellimages_rels_path).decode("utf-8")
        for m in re.finditer(r'Id="(rId\d+)"[^>]*Target="([^"]*)"', rels_content):
            rId = m.group(1)
            target = m.group(2)
            # cellimages的rels中Target可能是 media/xxx 或 ../media/xxx
            if target.startswith(".."):
                media_path = "xl/" + target.replace("../", "")
            elif target.startswith("xl/"):
                media_path = target
            else:
                media_path = f"xl/{target}"
            rId_to_media[rId] = media_path
    
    # Step 3: 解析 cellimages.xml → name到rId
    name_to_rId = {}
    cellimages_path = "xl/cellimages.xml"
    if cellimages_path in zf.namelist():
        ci_content = zf.read(cellimages_path).decode("utf-8")
        
        # 方法1: 正则分块（兼容命名空间变体）
        # 匹配 <etc:cellImage> ... </etc:cellImage> 或 <xdr:cellImage> ... </xdr:cellImage>
        blocks = re.findall(
            r'<\w+:cellImage[^>]*>.*?</\w+:cellImage>',
            ci_content, re.DOTALL
        )
        
        # 方法2: 如果方法1没匹配到，直接用正则从全文找name和r:embed的配对
        if not blocks:
            # 先按blipFill分块（每个cellImage至少有一个blipFill）
            blocks = re.split(r'(?=<\w+:cellImage)', ci_content)
            blocks = [b for b in blocks if 'cellImage' in b and 'blip' in b]
        
        for block in blocks:
            # 提取name（在cNvPr标签上）
            name_match = re.search(r'name="(ID_[^"]*)"', block)
            if not name_match:
                # 也尝试不带ID_前缀的
                name_match = re.search(r'cNvPr[^>]*name="([^"]*)"', block)
            if not name_match:
                continue
            name = name_match.group(1)
            
            # 提取r:embed (blip)
            embed_match = re.search(r'r:embed="(rId\d+)"', block)
            if embed_match:
                name_to_rId[name] = embed_match.group(1)
    
    # Step 4: 组合链路 → 读取图片
    images = []
    for dispimg_id, cell_ref in dispimg_cells.items():
        rId = name_to_rId.get(dispimg_id)
        if not rId:
            continue
        
        media_path = rId_to_media.get(rId)
        if not media_path or media_path not in zf.namelist():
            continue
        
        image_bytes = zf.read(media_path)
        fmt = Path(media_path).suffix.lstrip(".").lower()
        if fmt == "jpeg":
            fmt = "jpg"
        
        images.append({
            "cell": cell_ref,
            "type": "cell_embedded",
            "format": fmt,
            "bytes": image_bytes,
            "media_path": media_path,
            "dispimg_id": dispimg_id,
        })
    
    return images


# ============================================================
# 主函数
# ============================================================

def extract_sheet_images(
    file_path: str,
    sheet_name: str,
) -> list[dict]:
    """
    提取Excel某个sheet中的所有图片。
    
    Args:
        file_path: Excel文件路径
        sheet_name: sheet名称
    
    Returns:
        图片列表，每项包含:
        - cell: 单元格坐标 (如 "B5")
        - type: "floating" 或 "cell_embedded"
        - format: 图片格式 (如 "png", "jpg")
        - bytes: 图片二进制数据
        - media_path: zip内的原始路径
        - dispimg_id: (仅cell_embedded类型) DISPIMG的ID
    """
    file_path = str(file_path)
    if not Path(file_path).exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    with zipfile.ZipFile(file_path, "r") as zf:
        # 获取sheet映射
        sheet_map = _get_sheet_file_map(zf)
        
        if sheet_name not in sheet_map:
            available = list(sheet_map.keys())
            raise ValueError(f"Sheet '{sheet_name}' 不存在，可用sheet: {available}")
        
        sheet_info = sheet_map[sheet_name]
        
        # 提取两种类型的图片
        floating = _extract_floating_images(zf, sheet_info)
        embedded = _extract_cell_images(zf, sheet_info)
    
    # 合并结果，浮动图片放前面
    all_images = floating + embedded
    
    return all_images


def save_images(images: list[dict], output_dir: str, prefix: str = "") -> list[str]:
    """
    将提取的图片保存到文件。
    
    Args:
        images: extract_sheet_images()的返回值
        output_dir: 输出目录
        prefix: 文件名前缀
    
    Returns:
        保存的文件路径列表
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    saved = []
    for i, img in enumerate(images):
        cell = img["cell"]
        fmt = img["format"]
        img_type = img["type"]
        
        filename = f"{prefix}{'_' if prefix else ''}{cell}_{img_type}_{i}.{fmt}"
        filepath = output_dir / filename
        
        with open(filepath, "wb") as f:
            f.write(img["bytes"])
        
        saved.append(str(filepath))
    
    return saved


# ============================================================
# CLI 入口
# ============================================================

if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")
    
    if len(sys.argv) < 3:
        print("用法: python excel_image_extractor.py <xlsx文件> <sheet名称> [--save 目录]")
        print("示例: python excel_image_extractor.py data/excel/xxx.xlsx '4-竞品分析' --save output/images")
        sys.exit(1)
    
    xlsx_path = sys.argv[1]
    sheet = sys.argv[2]
    save_dir = None
    
    if "--save" in sys.argv:
        idx = sys.argv.index("--save")
        if idx + 1 < len(sys.argv):
            save_dir = sys.argv[idx + 1]
        else:
            save_dir = "output/images"
    
    print(f"📂 文件: {xlsx_path}")
    print(f"📋 Sheet: {sheet}")
    print()
    
    images = extract_sheet_images(xlsx_path, sheet)
    
    print(f"找到 {len(images)} 张图片:\n")
    
    for i, img in enumerate(images):
        size_kb = len(img["bytes"]) / 1024
        print(f"  [{i+1}] 单元格={img['cell']}  类型={img['type']}  格式={img['format']}  大小={size_kb:.1f}KB")
        if img.get("dispimg_id"):
            print(f"       DISPIMG_ID={img['dispimg_id']}")
    
    if save_dir and images:
        saved = save_images(images, save_dir)
        print(f"\n💾 已保存 {len(saved)} 张图片到 {save_dir}/")
        for s in saved:
            print(f"  → {s}")
