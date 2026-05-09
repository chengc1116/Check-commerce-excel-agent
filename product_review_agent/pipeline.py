# -*- coding: utf-8 -*-
"""
全流程异步编排器 — 产品立项审核 Pipeline

统一调度入口，将 Excel解析 → 图片提取 → 公共分析 → 专项分析 → 量化打分 → 报告整合 串联起来。

设计原则：
  1. 异步并行：Step1内部并行、Step2内部并行，最大化利用等待时间
  2. 解耦：每个环节独立可测试，pipeline只负责编排
  3. 降级：任何环节失败不阻塞其他环节，降级返回部分结果

用法：
    from product_review_agent.pipeline import run_pipeline

    result = await run_pipeline("xxx.xlsx", task_type="hot_upgrade")
    print(result.report)
"""

from __future__ import annotations

import os
from pathlib import Path as _Path
from dotenv import load_dotenv
load_dotenv(_Path(__file__).resolve().parent.parent / ".env", override=True)

import asyncio
import json
import logging
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


# ============================================================
# 结果数据类
# ============================================================

@dataclass
class PipelineResult:
    """Pipeline 执行结果"""
    file_name: str
    task_type: str = ""
    task_label: str = ""

    # Step 1: 解析结果
    project_data: dict = field(default_factory=dict)    # Excel解析JSON
    images: list = field(default_factory=list)           # 提取的图片
    parse_error: Optional[str] = None

    # Step 2: 分析结果
    common_analysis: dict = field(default_factory=dict)  # 公共分析（基本信息+人群+场景）
    common_scores: dict = field(default_factory=dict)    # 公共打分
    specific_analysis: dict = field(default_factory=dict)  # 专项分析结果
    specific_score: Optional[dict] = None               # 专项打分结果

    # Step 3: 综合评估
    overall_score: int = 0
    risk_level: str = "未知"
    report: str = ""

    # 元信息
    elapsed_seconds: float = 0.0
    error: Optional[str] = None

    def to_dict(self) -> dict:
        result = {
            "file_name": self.file_name,
            "task_type": self.task_type,
            "task_label": self.task_label,
            "overall_score": self.overall_score,
            "risk_level": self.risk_level,
            "elapsed_seconds": round(self.elapsed_seconds, 1),
            "common_scores": self.common_scores,
            "specific_analysis": self.specific_analysis,
            "specific_score": self.specif_score_to_dict(),
            "error": self.error,
        }
        # 如果specific_analysis包含项目数据，补充product_code和brand到顶层
        if self.specific_analysis:
            for key in ("product_code", "brand", "category_info", "project_data", "vl_report", "market_overview"):
                if key in self.specific_analysis and key not in result:
                    result[key] = self.specific_analysis[key]
        return result

    def specif_score_to_dict(self) -> dict:
        if self.specific_score and hasattr(self.specific_score, "to_dict"):
            return self.specific_score.to_dict()
        return self.specific_score or {}


# ============================================================
# Step 1: Excel 解析 + 图片提取（并行）
# ============================================================

def _save_parsed_json(file_stem: str, project_data: dict):
    """将 Excel 解析后的 JSON 保存到 output 目录"""
    try:
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        json_path = output_dir / f"{file_stem}_parsed_{timestamp}.json"

        # 过滤掉内部字段，只保留业务数据
        clean_data = {k: v for k, v in project_data.items() if not k.startswith("_")}

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(clean_data, f, ensure_ascii=False, indent=2)

        logger.info(f"[Pipeline] Excel解析JSON已保存: {json_path}")
    except Exception as e:
        logger.warning(f"[Pipeline] 保存解析JSON失败(不影响主流程): {e}")


async def _parse_excel(file_path: str, task_type: str = "") -> dict:
    """Excel → 结构化JSON"""
    try:
        from product_review_agent.parsers.excel_parsing_agent import parse_excel_to_project_review
        result = await parse_excel_to_project_review(file_path, task_type=task_type)

        if result.get("_error"):
            logger.error(f"[Pipeline] Excel解析出错: {result['_error']}")
            return {"_parse_error": result["_error"]}

        if result.get("_status") == "llm_unavailable":
            return {"_parse_error": "LLM不可用，无法解析Excel"}

        logger.info(f"[Pipeline] Excel解析完成, 字段数: {len([k for k in result if not k.startswith('_')])}")
        return result

    except Exception as e:
        logger.error(f"[Pipeline] Excel解析异常: {e}", exc_info=True)
        return {"_parse_error": str(e)}


def _extract_images(file_path: str) -> list[dict]:
    """提取Excel所有sheet的图片（同步，在线程池中执行）"""
    try:
        from product_review_agent.parsers.excel_image_extractor import extract_sheet_images
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        all_images = []
        for sheet_name in sheet_names:
            if sheet_name.startswith("~$"):
                continue
            try:
                images = extract_sheet_images(file_path, sheet_name)
                for img in images:
                    img["sheet_name"] = sheet_name
                all_images.extend(images)
            except Exception as e:
                logger.warning(f"[Pipeline] 提取sheet '{sheet_name}' 图片失败: {e}")

        logger.info(f"[Pipeline] 图片提取完成, 共 {len(all_images)} 张")
        return all_images

    except Exception as e:
        logger.error(f"[Pipeline] 图片提取异常: {e}", exc_info=True)
        return []


# ============================================================
# Step 2: 公共分析（基本信息 + 人群 + 场景）
# ============================================================

def _extract_basic_info(project_data: dict) -> dict:
    """从解析JSON中提取基本信息"""
    return {
        "product_name": project_data.get("project_name", ""),
        "brand": project_data.get("brand", ""),
        "category_l1": project_data.get("categoryl1", ""),
        "category_l2": project_data.get("categoryl2", ""),
        "category_l3": project_data.get("categoryl3", ""),
        "applicant": project_data.get("applicant", ""),
        "project_time": project_data.get("project_time", ""),
        "design_time": project_data.get("design_time", ""),
        "proofing_time": project_data.get("proofing_time", ""),
        "launch_time": project_data.get("launch_time", ""),
        "market_size": project_data.get("market_size", ""),
        "estimated_sales": project_data.get("estimated_sales", ""),
        "pricing": project_data.get("pricing", ""),
        "gfm": project_data.get("gfm", ""),
        "ERP_price": project_data.get("ERP_price", ""),
        "core_config": project_data.get("core_config", ""),
        "design_require": project_data.get("design_require") or {"content": project_data.get("design_purpose", "")},
        "product_comparison": project_data.get("product_comparison") or {"competitor_price": project_data.get("competitor_price", "")},
    }


async def _score_audience_scenario(project_data: dict) -> dict:
    """人群+场景合并分析评分"""
    from product_review_agent.reviewer import ascore_with_llm, _fallback_score, _build_audience_scenario_prompt
    from product_review_agent.agents.llm_client import get_llm_client

    llm = get_llm_client()

    # 提取人群文本
    used_people = project_data.get("used_people", [])
    if used_people:
        audience_text = json.dumps(used_people, ensure_ascii=False, indent=2)
    else:
        audience_text = project_data.get("people_analysis", "")

    # 提取场景文本
    used_scene = project_data.get("used_scene", [])
    if used_scene:
        scenario_text = json.dumps(used_scene, ensure_ascii=False, indent=2)
    else:
        scenario_text = project_data.get("scene_analysis", "") or project_data.get("usage_scenarios", "")

    # 两者都为空则回退
    has_audience = bool(audience_text) and not ("[图片" in audience_text and len(audience_text.strip()) < 50)
    has_scenario = bool(scenario_text) and not ("[图片" in scenario_text and len(scenario_text.strip()) < 50)

    if not has_audience and not has_scenario:
        return _fallback_score("audience_scenario", reason="no_data")

    # 构建产品上下文
    product_context = {
        "product_name": project_data.get("project_name") or project_data.get("product_name", "未知"),
        "brand": project_data.get("brand", "未知"),
        "category_l1": project_data.get("category_l1") or project_data.get("categoryl1", ""),
        "category_l2": project_data.get("category_l2") or project_data.get("categoryl2", ""),
        "category_l3": project_data.get("category_l3") or project_data.get("categoryl3", ""),
        "pricing": project_data.get("pricing", "未填写"),
        "estimated_sales": project_data.get("estimated_sales", "未填写"),
    }

    # 构建完整prompt
    prompt = _build_audience_scenario_prompt(audience_text, scenario_text, product_context)

    return await ascore_with_llm(llm, prompt, "audience_scenario")


async def _analyze_common(project_data: dict) -> dict:
    """
    公共分析：基本信息 + 人群场景合并评分
    """
    # 基本信息（同步提取）
    basic_info = _extract_basic_info(project_data)

    # 人群+场景合并评分
    audience_scenario_score = await _score_audience_scenario(project_data)

    return {
        "basic_info": basic_info,
        "audience_scenario_score": audience_scenario_score,
    }


# ============================================================
# Step 3: 专项分析（根据任务类型选择分析器）
# ============================================================

def _get_analyzer(task_type: str):
    """根据任务类型获取对应分析器实例"""
    from product_review_agent.analyzers.hot_upgrade_analyzer import HotUpgradeAnalyzer
    from product_review_agent.analyzers.competitor_upgrade_analyzer import CompetitorUpgradeAnalyzer
    from product_review_agent.analyzers.low_sale_iterate_analyzer import LowSaleIterateAnalyzer
    from product_review_agent.analyzers.category_gap_analyzer import CategoryGapAnalyzer

    analyzers = {
        "hot_upgrade": HotUpgradeAnalyzer,
        "competitor_upgrade": CompetitorUpgradeAnalyzer,
        "low_sale_iterate": LowSaleIterateAnalyzer,
        "category_gap": CategoryGapAnalyzer,
    }

    cls = analyzers.get(task_type)
    if not cls:
        logger.error(f"[Pipeline] 未知任务类型: {task_type}")
        return None

    return cls()


async def _analyze_specific(task_type: str, project_data: dict, images: list) -> tuple[dict, Optional[object]]:
    """执行专项分析 + 量化打分"""
    analyzer = _get_analyzer(task_type)
    if not analyzer:
        return {}, None

    try:
        analysis_result, score_result = await analyzer.run(project_data, images)
        return analysis_result, score_result
    except Exception as e:
        logger.error(f"[Pipeline] 专项分析异常({task_type}): {e}", exc_info=True)
        return {"_error": str(e)}, None


# ============================================================
# Step 4: 同类产品分析（与专项分析并行）
# ============================================================

async def _analyze_product_history(project_data: dict) -> dict:
    """同类产品+销量分析"""
    try:
        from product_review_agent.reviewer import analyze_with_history
        return await analyze_with_history(project_data)
    except Exception as e:
        logger.error(f"[Pipeline] 同类产品分析异常: {e}", exc_info=True)
        return {"products": [], "analysis": None, "error": str(e)}


# ============================================================
# Step 5: 报告整合
# ============================================================

SEPARATOR = "=" * 78
THIN_SEP = "-" * 78


def _stars_display(n: int) -> str:
    return "★" * n + "☆" * (5 - n)


def _score_stars(score: int) -> int:
    if score >= 90: return 5
    elif score >= 75: return 4
    elif score >= 60: return 3
    elif score >= 40: return 2
    else: return 1


def _append_score_detail(lines: list, score: dict, total: int, section_name: str):
    """将评分明细追加到报告行"""
    if score and total > 0:
        lines.append("[评分明细]")
        for dim_name, dim_info in score.get("dimensions", {}).items():
            s = dim_info.get("score", 0) if isinstance(dim_info, dict) else 0
            r = dim_info.get("reason", "") if isinstance(dim_info, dict) else str(dim_info)
            lines.append(f"  {dim_name}: {s}/25 - {r}")
        lines.append("")

        for label, key in [("优势", "strengths"), ("不足", "weaknesses"), ("改进建议", "suggestions")]:
            items = score.get(key, [])
            if items:
                prefix = "+" if label == "优势" else ("-" if label == "不足" else ">")
                lines.append(f"[{label}]")
                for item in items:
                    lines.append(f"  {prefix} {item}")
                lines.append("")
    else:
        lines.append(f"  ({section_name}评分不可用)")
        lines.append("")


def _generate_report(
    project_data: dict,
    common_analysis: dict,
    specific_analysis: dict,
    specific_score,
    product_analysis: dict,
    task_type: str,
    task_label: str,
) -> str:
    """生成最终整合报告"""
    from product_review_agent.analyzers.base import AnalyzerScore

    lines = []

    # 标题
    lines.append(SEPARATOR)
    lines.append("                    产品立项审核报告")
    lines.append(SEPARATOR)
    lines.append("")

    # 产品概览
    basic = common_analysis.get("basic_info", {})
    product_name = basic.get("product_name", "(未填写)")
    brand = basic.get("brand", "(未填写)")
    cat = " > ".join(filter(None, [
        basic.get("category_l1", ""),
        basic.get("category_l2", ""),
        basic.get("category_l3", ""),
    ])) or "(未填写)"
    owner = basic.get("applicant", "(未填写)")

    lines.append(f"产品名称: {product_name}")
    lines.append(f"品牌: {brand}")
    lines.append(f"品类: {cat}")
    lines.append(f"负责人: {owner}")
    lines.append(f"审核类型: {task_label}")
    lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")

    # 一、立项信息
    lines.append(THIN_SEP)
    lines.append("一、立项信息")
    lines.append(THIN_SEP)
    for key, label in [
        ("project_time", "立项时间"), ("design_time", "设计时间"),
        ("proofing_time", "打样时间"), ("launch_time", "上架时间"),
    ]:
        lines.append(f"  {label}: {basic.get(key, '(未填写)')}")
    lines.append("")

    # 二、市场与定价
    lines.append(THIN_SEP)
    lines.append("二、市场与定价")
    lines.append(THIN_SEP)
    lines.append(f"  市场规模: {basic.get('market_size', '(未填写)')}")
    lines.append(f"  目标销量: {basic.get('estimated_sales', '(未填写)')}")
    lines.append(f"  定价: {basic.get('pricing', '(未填写)')}")
    lines.append(f"  毛利率: {basic.get('gfm', '(未填写)')}")
    lines.append(f"  ERP成本: {basic.get('ERP_price', '(未填写)')}")
    lines.append("")

    # 竞品对比信息（LLM可能返回dict或list）
    comparison = basic.get("product_comparison", {})
    if comparison:
        # 兼容 list 和 dict 两种格式
        if isinstance(comparison, list):
            for i, comp in enumerate(comparison):
                prefix = f"  竞品{i + 1} " if len(comparison) > 1 else "  "
                if isinstance(comp, dict):
                    lines.append(f"{prefix}对手商品: {comp.get('comparison_name', '(未填写)')}")
                    lines.append(f"{prefix}对手卖点(复制): {comp.get('selling_point', '(未填写)')}")
                    lines.append(f"{prefix}对手卖点(超越): {comp.get('improving_point', '(未填写)')}")
                else:
                    lines.append(f"{prefix}{comp}")
            lines.append("")
        else:
            lines.append(f"  对手商品: {comparison.get('comparison_name', '(未填写)')}")
            lines.append(f"  对手卖点(复制): {comparison.get('selling_point', '(未填写)')}")
            lines.append(f"  对手卖点(超越): {comparison.get('improving_point', '(未填写)')}")
            lines.append("")

    # 三、人群与场景分析（合并）
    as_score = common_analysis.get("audience_scenario_score", {})
    as_total = as_score.get("total_score", 0)

    lines.append(THIN_SEP)
    lines.append(f"三、人群与场景分析 [评分: {as_total}/100] [{_stars_display(_score_stars(as_total))}]")
    lines.append(THIN_SEP)

    # 分析段落
    analysis = as_score.get("analysis", {})
    if analysis:
        dim_labels = {
            "audience_scene_fit": "人群-场景匹配度",
            "insight_depth": "需求洞察深度",
            "data_coverage": "数据支撑与覆盖度",
            "commercial_value": "商业价值判断",
        }
        for key, label in dim_labels.items():
            text = analysis.get(key, "")
            if text:
                lines.append(f"  [{label}]")
                for para_line in text.split("\n"):
                    lines.append(f"    {para_line}")
                lines.append("")

    # 专家分析
    expert = as_score.get("expert_analysis", {})
    if expert:
        lines.append("  ── 专家视角：人群与场景建议 ──")
        if expert.get("target_audience"):
            lines.append(f"  建议目标人群: {expert['target_audience']}")
        if expert.get("core_scenarios"):
            lines.append(f"  建议核心场景: {expert['core_scenarios']}")
        if expert.get("key_suggestion"):
            lines.append(f"  关键建议: {expert['key_suggestion']}")
        lines.append("")

    # 评分明细
    scores = as_score.get("scores", {})
    if scores:
        lines.append("  [评分明细]")
        for dim_name, dim_info in scores.items():
            if isinstance(dim_info, dict):
                s = dim_info.get("score", 0)
                r = dim_info.get("reason", "")
                lines.append(f"    {dim_name}: {s}/25 - {r}")
        lines.append("")

    # 优劣势和建议
    for label, key in [("优势", "strengths"), ("不足", "weaknesses"), ("改进建议", "suggestions")]:
        items = as_score.get(key, [])
        if items:
            prefix = "+" if label == "优势" else ("-" if label == "不足" else ">")
            lines.append(f"  [{label}]")
            for item in items:
                lines.append(f"    {prefix} {item}")
            lines.append("")
    if not analysis and not expert:
        _append_score_detail(lines, as_score, as_total, "人群与场景")

    # 四、专项分析（核心部分）
    if specific_analysis and not specific_analysis.get("_error"):
        analyzer = _get_analyzer(task_type)
        if analyzer:
            specific_report = analyzer.format_report(specific_analysis, specific_score)
            lines.append(THIN_SEP)
            lines.append(f"四、{task_label}专项分析")
            lines.append(THIN_SEP)
            # 缩进专项报告
            for line in specific_report.split("\n"):
                lines.append(f"  {line}")
            lines.append("")
        else:
            lines.append(THIN_SEP)
            lines.append(f"四、{task_label}专项分析 (分析器不可用)")
            lines.append(THIN_SEP)
            lines.append("")
    else:
        lines.append(THIN_SEP)
        lines.append("四、专项分析 (不可用)")
        lines.append(THIN_SEP)
        err = specific_analysis.get("_error", "未知原因") if specific_analysis else "无分析结果"
        lines.append(f"  专项分析不可用: {err}")
        lines.append("")

    # 五、同类产品及销售情况
    if product_analysis and product_analysis.get("products"):
        products = product_analysis["products"]
        analysis = product_analysis.get("analysis")
        category_l2 = basic.get("category_l2", "未知")

        lines.append(THIN_SEP)
        lines.append("五、同类产品及销售情况")
        lines.append(THIN_SEP)
        lines.append(f"  品类「{category_l2}」现有活跃产品（共 {len(products)} 个）:")
        lines.append("")

        # 产品数据表格
        lines.append("  %-12s %-10s %-8s %-16s %-14s %-14s %s" % (
            "货号", "品牌", "版本", "三级品类", "销量(较早)", "销量(较近)", "趋势"))
        lines.append("  " + "-" * 90)

        for p in products:
            sku = p.get("sku", "-")
            pbrand = p.get("brand", "-")
            version = p.get("version", "-")
            cat3 = p.get("category_l3", "-")
            if len(cat3) > 14:
                cat3 = cat3[:12] + ".."
            sales = p.get("recent_sales", [])

            if len(sales) >= 2:
                m1_val = sales[1].get("sales_volume", 0)
                m2_val = sales[0].get("sales_volume", 0)
                m1_label = sales[1].get("month", "?")
                m2_label = sales[0].get("month", "?")
                if m2_val > m1_val:
                    trend = "↑"
                elif m2_val < m1_val:
                    trend = "↓"
                else:
                    trend = "—"
                s1 = "%s:%d" % (m1_label[5:], m1_val)
                s2 = "%s:%d" % (m2_label[5:], m2_val)
            elif len(sales) == 1:
                s1 = "-"
                m_val = sales[0].get("sales_volume", 0)
                m_label = sales[0].get("month", "?")
                s2 = "%s:%d" % (m_label[5:], m_val)
                trend = ""
            else:
                s1 = "-"
                s2 = "暂无"
                trend = ""

            lines.append("  %-12s %-10s %-8s %-16s %-14s %-14s %s" % (
                sku[:10], pbrand[:8], version[:6], cat3, s1, s2, trend))

            # 显示 CBB 模块明细
            modules = p.get("modules", [])
            if modules:
                module_brief = ", ".join(
                    f"{m.get('cbb_code', '')}({m.get('cbb_name', '')})"
                    for m in modules[:5]
                )
                if len(modules) > 5:
                    module_brief += f" 等{len(modules)}个"
                lines.append(f"    模块: {module_brief}")

        lines.append("")

        # AI 分析
        if analysis:
            lines.append(THIN_SEP)
            lines.append("  【AI 分析建议】")
            lines.append(THIN_SEP)
            if analysis.get("analysis"):
                for al in analysis["analysis"].split("\n"):
                    lines.append(f"  {al}")
                lines.append("")
            if analysis.get("suggestions"):
                for sg in analysis["suggestions"]:
                    lines.append(f"  > {sg}")
                lines.append("")

    # 六、综合评估
    lines.append(THIN_SEP)
    lines.append("六、综合评估")
    lines.append(THIN_SEP)

    # 综合分 = 专项分析得分（人群场景仅做分析，不纳入评分）
    specific_total = 0
    if specific_score and isinstance(specific_score, object) and hasattr(specific_score, "total_score"):
        specific_total = specific_score.total_score
    elif specific_score and isinstance(specific_score, dict):
        specific_total = specific_score.get("total_score", 0)

    overall = specific_total
    risk = "未知"
    if overall > 0:
        lines.append(f"  专项评分: {overall}/100")
        lines.append(f"  综合评分: {overall}/100")
        lines.append(f"  星级: [{_stars_display(_score_stars(overall))}]")

        if overall >= 75:
            risk = "低"
        elif overall >= 50:
            risk = "中"
        else:
            risk = "高"
        lines.append(f"  风险等级: {risk}")
    else:
        lines.append("  综合评分: 暂无法评估（缺少评分数据或LLM不可用）")

    lines.append("")

    # 设计要求
    design_req = basic.get("design_require", {})
    if not design_req and project_data.get("design_purpose"):
        design_req = {"content": project_data["design_purpose"]}
        if project_data.get("upgrade_modules"):
            design_req["upgrade_modules"] = project_data["upgrade_modules"]
        if project_data.get("upgrade_valiable"):
            design_req["upgrade_valiable"] = project_data["upgrade_valiable"]
    if design_req:
        lines.append(THIN_SEP)
        lines.append("附、设计要求")
        lines.append(THIN_SEP)
        if design_req.get("content"):
            lines.append(f"  设计目的: {design_req['content']}")
        if design_req.get("upgrade_modules"):
            lines.append(f"  升级模块: {design_req['upgrade_modules']}")
        if design_req.get("upgrade_valiable"):
            lines.append(f"  升级可行性: {design_req['upgrade_valiable']}")
        lines.append("")

    lines.append(SEPARATOR)
    lines.append("报告结束")
    lines.append(SEPARATOR)

    return "\n".join(lines)


# ============================================================
# 核心编排函数
# ============================================================

async def run_pipeline(file_path: str, task_type: str) -> PipelineResult:
    """
    全流程异步编排。

    Args:
        file_path: Excel文件路径
        task_type: 任务类型 (hot_upgrade/competitor_upgrade/low_sale_iterate/category_gap)

    Returns:
        PipelineResult
    """
    from product_review_agent.feishu.session_manager import TASK_TYPE_MAP
    from product_review_agent.product_db.operation_logger import OperationLogger

    start_time = time.time()
    file_path = Path(file_path)
    task_obj = TASK_TYPE_MAP.get(task_type)
    task_label = task_obj.label if task_obj else task_type
    task_emoji = task_obj.emoji if task_obj else "📋"

    # 操作日志记录器
    op_log = OperationLogger()

    logger.info(f"[Pipeline] ========== 审核流水线启动 ==========")
    logger.info(f"[Pipeline] 文件: {file_path.name}, 类型: {task_emoji} {task_label}")

    # 记录：审核启动
    op_log.log("pipeline_start", "review", target=file_path.name,
               detail=f"{task_emoji} {task_label}", extra={"task_type": task_type})

    if not file_path.exists():
        op_log.log("pipeline_start", "review", target=file_path.name,
                   detail="文件不存在", status="failed")
        op_log.close()
        return PipelineResult(
            file_name=file_path.name,
            task_type=task_type,
            task_label=task_label,
            error=f"文件不存在: {file_path}",
        )

    # ==================== Step 1: Excel解析 + 图片提取（并行） ====================
    logger.info("[Pipeline] Step 1: Excel解析 + 图片提取...")
    step1_start = time.time()

    # 图片提取是同步IO密集型，放到线程池
    loop = asyncio.get_event_loop()
    project_data, images = await asyncio.gather(
        _parse_excel(str(file_path), task_type=task_type),
        loop.run_in_executor(None, _extract_images, str(file_path)),
    )

    parse_error = project_data.get("_parse_error")
    if parse_error:
        logger.error(f"[Pipeline] Excel解析失败: {parse_error}")
        # 解析失败但有图片，降级：用空数据继续
        project_data_clean = {}
    else:
        project_data_clean = {k: v for k, v in project_data.items() if not k.startswith("_")}

    # 保存 Excel 解析 JSON 到 output 目录
    _save_parsed_json(file_path.stem, project_data)

    # 记录：Excel解析结果
    step1_ms = int((time.time() - step1_start) * 1000)
    field_count = len(project_data_clean)
    op_log.log("excel_parse", "review", target=file_path.name,
               detail=f"解析{'失败' if parse_error else '完成'}, {field_count}个字段, {len(images)}张图片",
               status="failed" if parse_error else "success",
               elapsed_ms=step1_ms,
               extra={"fields": field_count, "images": len(images), "parse_error": parse_error})

    # ==================== Step 2: 公共分析 + 专项分析 + 同类产品分析（并行） ====================
    logger.info("[Pipeline] Step 2: 并行分析 (公共+专项+同类产品)...")
    step2_start = time.time()

    # 准备图片数据给专项分析
    # 只传递图片bytes（用于VL模型），不需要元数据
    image_bytes_list = []
    for img in images:
        if img.get("bytes"):
            image_bytes_list.append(img["bytes"])

    common_analysis, (specific_analysis, specific_score), product_analysis = await asyncio.gather(
        _analyze_common(project_data_clean),
        _analyze_specific(task_type, project_data_clean, image_bytes_list),
        _analyze_product_history(project_data_clean),
    )

    logger.info(f"[Pipeline] 分析完成: 公共={bool(common_analysis)}, 专项={bool(specific_analysis)}, 同类={bool(product_analysis)}")

    # 记录：分析阶段结果
    step2_ms = int((time.time() - step2_start) * 1000)
    as_total = common_analysis.get("audience_scenario_score", {}).get("total_score", 0)
    specific_total = 0
    if specific_score and hasattr(specific_score, "total_score"):
        specific_total = specific_score.total_score

    op_log.log("analysis_complete", "review", target=file_path.name,
               detail=f"人群场景{as_total}/专项{specific_total}",
               elapsed_ms=step2_ms,
               extra={"audience_scenario_score": as_total,
                      "specific_score": specific_total, "task_type": task_type})

    # ==================== Step 3: 报告整合 ====================
    logger.info("[Pipeline] Step 3: 报告整合...")

    report = _generate_report(
        project_data=project_data_clean,
        common_analysis=common_analysis,
        specific_analysis=specific_analysis,
        specific_score=specific_score,
        product_analysis=product_analysis,
        task_type=task_type,
        task_label=f"{task_emoji} {task_label}",
    )

    # 综合分 = 专项分析得分（人群场景仅做分析，不纳入评分）
    valid_scores = []
    if specific_total > 0:
        valid_scores.append(("专项", specific_total, 1.0))

    overall = 0
    risk = "未知"
    if valid_scores:
        total_weight = sum(w for _, _, w in valid_scores)
        if total_weight > 0:
            overall = int(round(sum(s * w / total_weight for _, s, w in valid_scores)))
        if overall >= 75:
            risk = "低"
        elif overall >= 50:
            risk = "中"
        else:
            risk = "高"

    # 构建公共评分dict（供飞书卡片使用）
    common_scores = {
        "audience_scenario": common_analysis.get("audience_scenario_score", {}),
    }

    elapsed = time.time() - start_time
    elapsed_ms = int(elapsed * 1000)
    logger.info(f"[Pipeline] ========== 完成, 综合分: {overall}/100, 风险: {risk}, 耗时: {elapsed:.1f}s ==========")

    # 记录：Pipeline完成
    op_log.log("pipeline_complete", "review", target=file_path.name,
               detail=f"综合分{overall}/100, 风险{risk}, {task_emoji}{task_label}",
               elapsed_ms=elapsed_ms,
               extra={"overall_score": overall, "risk_level": risk,
                      "task_type": task_type, "parse_error": parse_error is not None})

    op_log.close()

    return PipelineResult(
        file_name=file_path.name,
        task_type=task_type,
        task_label=f"{task_emoji} {task_label}",
        project_data=project_data_clean,
        images=images,
        parse_error=parse_error,
        common_analysis=common_analysis,
        common_scores=common_scores,
        specific_analysis=specific_analysis,
        specific_score=specific_score,
        overall_score=overall,
        risk_level=risk,
        report=report,
        elapsed_seconds=elapsed,
    )
