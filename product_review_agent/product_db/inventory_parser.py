# -*- coding: utf-8 -*-
"""
货盘表解析器 v2 — 解析公司月度货盘 Excel

v2 变更:
  - 动态检测品牌列（从 header 行读取，兼容不同品牌布局）
  - 货号不再拆分（"H66/H66JYD/H69/H69JYD" 整体保留）
  - 图片命名为 {品类l1}_{sku}.ext，路径存入数据库
  - 支持 WPS DISPIMG 函数嵌入的图片（通过 cellimages.xml 映射）
  - 支持三种不同格式的货盘表

表格格式特点:
  - 多层合并单元格（一级品类A、二级品类B、三级品类C、版本D）
  - 每个产品占3-5行: 图片行 → 货号行 → 销量行 → [升级情况行] → [空行]
  - 品牌分布在不同列（需要从 header 动态检测）
  - 销量格式: "三月：2819-452=2367+206=2573" (最后一个数字为最终销量)
  - 图片使用 WPS 的 DISPIMG 函数嵌入

三种表格差异:
  - 护膝: F+G=TMT(合并), H+I=SERUNA(合并), J=JAFFICK, K=ANTA
  - 护腕: F=TMT, G=SERUNA, H=JAFFICK, I=ANTA
  - 护踝: F=TMT, G=SERUNA, H=JAFFICK, I=ANTA
"""

from __future__ import annotations

import logging
import os
import re
import zipfile
from dataclasses import dataclass
from typing import Optional

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class ParsedProduct:
    """解析出的单个产品记录"""
    category_l1: str = ""
    category_l2: str = ""
    category_l3: str = ""
    version: str = ""
    brand: str = ""
    sku: str = ""
    sales_volume: int = 0
    status: str = "active"
    image_path: Optional[str] = None
    raw_sales_text: str = ""
    month: str = ""  # 推断的月份
    # 内部: 图片源文件路径（用于后续重命名复制）
    _image_source: Optional[str] = None

    def to_dict(self) -> dict:
        return {
            "category_l1": self.category_l1,
            "category_l2": self.category_l2,
            "category_l3": self.category_l3,
            "version": self.version,
            "brand": self.brand,
            "sku": self.sku,
            "sales_volume": self.sales_volume,
            "status": self.status,
            "image_path": self.image_path,
        }


# ============================================================
# 销量解析
# ============================================================

def parse_sales_volume(text: str) -> int:
    """
    从销量文本中提取最终销量。
    支持: "三月：2819-452=2367+206=2573" → 2573, "266" → 266, "大货阶段" → 0
    """
    if not text or not text.strip():
        return 0
    text = text.strip()
    if text in ("0", "大货阶段", "已清仓", "清仓中"):
        return 0
    cleaned = text.replace("\n", " ").replace("\r", "")
    eq_matches = re.findall(r'=(-?\d+)', cleaned)
    if eq_matches:
        return max(int(eq_matches[-1]), 0)
    numbers = re.findall(r'(\d+)', cleaned)
    if numbers:
        return int(numbers[-1])
    return 0


def infer_status_from_text(text: str) -> str:
    """从升级情况文本推断产品状态"""
    if not text:
        return "active"
    if "清仓" in text:
        return "archived"
    return "active"


# ============================================================
# 图片提取
# ============================================================

def extract_images_from_xlsx(xlsx_path: str, output_dir: str) -> dict[str, str]:
    """从 xlsx 中提取图片文件。Returns: {原始media文件名: 绝对路径}"""
    os.makedirs(output_dir, exist_ok=True)
    result = {}
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        for mf in z.namelist():
            if not mf.startswith("xl/media/") or mf.endswith("/"):
                continue
            data = z.read(mf)
            fname = os.path.basename(mf)
            if not fname:
                nums = re.findall(r'\d+', mf)
                ext = os.path.splitext(mf)[1] or '.png'
                fname = f"image{nums[-1] if nums else '0'}{ext}"
            out_path = os.path.join(output_dir, fname)
            with open(out_path, 'wb') as f:
                f.write(data)
            result[fname] = out_path
    return result


def get_image_ext(filename: str) -> str:
    _, ext = os.path.splitext(filename)
    return ext if ext else '.png'


def sanitize_filename(name: str) -> str:
    """清理文件名中的非法字符"""
    name = name.replace("\n", "").replace("\r", "")
    name = name.replace("/", "_").replace("\\", "_")
    name = name.replace(":", "_").replace("*", "_")
    name = name.replace("?", "_").replace('"', "_")
    name = name.replace("<", "_").replace(">", "_")
    name = name.replace("|", "_")
    return name.strip()


# ============================================================
# 品牌列检测
# ============================================================

def detect_brand_columns(ws, header_rows: int = 3) -> dict[int, str]:
    """
    从 header 行动态检测品牌列映射。
    支持合并单元格（如护膝的 F+G=TMT, H+I=SERUNA）。
    Returns: {列号: 品牌名}
    """
    brand_names = {"TMT", "SERUNA", "JAFFICK", "ANTA"}
    col_brand = {}

    for mc in ws.merged_cells.ranges:
        if mc.min_row > header_rows:
            continue
        val = ws.cell(row=mc.min_row, column=mc.min_col).value
        if val and str(val).strip().upper() in brand_names:
            brand = str(val).strip().upper()
            for c in range(mc.min_col, mc.max_col + 1):
                col_brand[c] = brand

    for r in range(1, header_rows + 1):
        for c in range(1, ws.max_column + 1):
            if c in col_brand:
                continue
            val = ws.cell(row=r, column=c).value
            if val and str(val).strip().upper() in brand_names:
                col_brand[c] = str(val).strip().upper()

    return col_brand


def detect_data_start_row(ws) -> int:
    """检测数据开始的行号"""
    for r in range(1, min(10, ws.max_row + 1)):
        e_val = ws.cell(row=r, column=5).value
        if e_val and str(e_val).strip() == "图片":
            return r
    return 3


# ============================================================
# DISPIMG 图片映射 (WPS 特有)
# ============================================================

def build_dispimg_map(xlsx_path: str, sheet_name: str | None = None) -> dict[tuple[int, int], str]:
    """
    构建 {(行号, 列号): media文件名} 映射。
    通过解析 worksheet XML 中的 DISPIMG 函数 + cellimages.xml 映射。

    步骤:
      1. 从 workbook.xml + rels 找到目标 sheet 对应的 XML 文件
      2. 从 worksheet XML 提取所有 DISPIMG("ID_xxx") 的单元格位置和 ID
      3. 从 cellimages.xml + rels 建立 ID → media 文件名映射
      4. 合并得到 (行,列) → media 映射
    """
    result = {}

    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            # --- Step 1: 找 sheet 文件 ---
            wb_content = z.read('xl/workbook.xml').decode('utf-8', errors='ignore')
            sheets = re.findall(r'<sheet name="([^"]+)"[^>]*?r:id="(rId\d+)"', wb_content)

            wb_rels = z.read('xl/_rels/workbook.xml.rels').decode('utf-8', errors='ignore')
            rid_to_file = {}
            for m in re.finditer(r'Id="(rId\d+)".*?Target="([^"]+)"', wb_rels):
                rid_to_file[m.group(1)] = m.group(2)

            # 确定 target sheet
            target_srid = None
            if sheet_name:
                for sname, srid in sheets:
                    if sname == sheet_name:
                        target_srid = srid
                        break
            else:
                # 取最后一个非目的/TOP5的sheet
                for sname, srid in reversed(sheets):
                    if sname not in ('目的', 'TOP5'):
                        target_srid = srid
                        break

            if not target_srid or target_srid not in rid_to_file:
                logger.warning("未找到目标 sheet")
                return result

            sheet_file = 'xl/' + rid_to_file[target_srid]

            # --- Step 2: 提取 DISPIMG 位置和 ID ---
            sheet_content = z.read(sheet_file).decode('utf-8', errors='ignore')

            # 匹配: <c r="F8" ...><f>_xlfn.DISPIMG("ID_xxx",1)</f>...<v>...</v></c>
            # 注意: [^<]* 确保不跨越 </c> 标签（避免跨单元格匹配）
            dispimg_pattern = re.compile(
                r'<c r="([A-Z]+)(\d+)"[^>]*>[^<]*'
                r'(?:<[^/][^>]*>[^<]*)*?'
                r'DISPIMG\(&quot;([^&]+)&quot;',
                re.DOTALL
            )
            cell_id_map = {}  # {ID: (row, col)}
            for m in dispimg_pattern.finditer(sheet_content):
                col_str = m.group(1)
                row = int(m.group(2))
                img_id = m.group(3)
                # 列字母转数字 (A=1, B=2, ..., F=6)
                col = 0
                for ch in col_str:
                    col = col * 26 + (ord(ch) - ord('A') + 1)
                cell_id_map[img_id] = (row, col)

            logger.debug(f"DISPIMG cells found: {len(cell_id_map)}")

            # --- Step 3: cellimages.xml ID → rId → media ---
            if 'xl/cellimages.xml' not in z.namelist():
                logger.warning("cellimages.xml 不存在")
                return result

            ci_rels_content = z.read('xl/_rels/cellimages.xml.rels').decode('utf-8', errors='ignore')
            ci_rid_to_media = {}
            for m in re.finditer(r'Id="(rId\d+)".*?Target="([^"]+)"', ci_rels_content):
                ci_rid_to_media[m.group(1)] = os.path.basename(m.group(2))

            ci_content = z.read('xl/cellimages.xml').decode('utf-8', errors='ignore')

            # 匹配 name="ID_xxx" 和 r:embed="rIdN"
            id_to_media = {}
            for m in re.finditer(
                r'<xdr:pic>.*?name="([^"]*?ID_[^"]*?)"[^>]*?/>.*?r:embed="(rId\d+)"',
                ci_content, re.DOTALL
            ):
                pic_name = m.group(1)
                rid = m.group(2)
                media = ci_rid_to_media.get(rid)
                if media:
                    id_to_media[pic_name] = media

            # --- Step 4: 合并 ---
            for img_id, (row, col) in cell_id_map.items():
                # 精确匹配
                if img_id in id_to_media:
                    result[(row, col)] = id_to_media[img_id]
                else:
                    # 模糊匹配（ID可能截断）
                    for full_id, media in id_to_media.items():
                        if img_id in full_id or full_id in img_id:
                            result[(row, col)] = media
                            break

    except Exception as e:
        logger.warning(f"构建 DISPIMG 映射失败: {e}")

    logger.info(f"DISPIMG 映射: {len(result)} 个单元格→图片")
    return result


# ============================================================
# InventoryParser v2
# ============================================================

class InventoryParser:
    """货盘表解析器 v2"""

    def parse(
        self,
        file_path: str,
        sheet_name: str | None = None,
        month: str | None = None,
        extract_images: bool = True,
        images_dir: str | None = None,
        category_l1_override: str | None = None,
    ) -> list[ParsedProduct]:
        """解析货盘 Excel 文件"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # 选择 sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' 不存在，可选: {wb.sheetnames}")
            ws = wb[sheet_name]
        else:
            data_sheets = [n for n in wb.sheetnames if n not in ("目的", "TOP5")]
            if not data_sheets:
                raise ValueError("未找到数据 sheet")
            ws = wb[data_sheets[-1]]
            sheet_name = ws.title
            logger.info(f"自动选择 sheet: {sheet_name}")

        # 推断月份
        if not month:
            month = self._infer_month(sheet_name)
            logger.info(f"推断月份: {month}")

        # 动态检测品牌列
        brand_columns = detect_brand_columns(ws)
        logger.info(f"检测到品牌列: {brand_columns}")

        # 检测数据起始行
        data_start = detect_data_start_row(ws)
        logger.info(f"数据起始行: {data_start}")

        # 提取图片到临时目录（避免多文件冲突）
        img_files = {}
        target_images_dir = images_dir or os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(file_path))),
            "data", "images"
        )
        tmp_images_dir = None
        if extract_images:
            import tempfile
            tmp_images_dir = tempfile.mkdtemp(prefix="inv_imgs_")
            img_files = extract_images_from_xlsx(file_path, tmp_images_dir)

        # 构建 DISPIMG 图片位置映射
        dispimg_map = build_dispimg_map(file_path, sheet_name)
        logger.info(f"DISPIMG 映射条目: {len(dispimg_map)}")

        # 解析数据行
        products = self._parse_sheet(
            ws, month, brand_columns, dispimg_map, img_files,
            data_start, category_l1_override, target_images_dir,
        )

        # 清理临时图片目录
        if tmp_images_dir:
            import shutil
            try:
                shutil.rmtree(tmp_images_dir, ignore_errors=True)
            except Exception:
                pass

        wb.close()
        logger.info(f"解析完成: sheet={sheet_name}, month={month}, products={len(products)}")
        return products

    def _infer_month(self, sheet_name: str) -> str:
        """从 sheet 名称推断月份"""
        m = re.match(r'(\d{4})[.\-](\d{2})', sheet_name)
        if m:
            return f"{m.group(1)}-{m.group(2)}"
        m = re.match(r'(\d{2})年(\d{2})月', sheet_name)
        if m:
            return f"20{m.group(1)}-{m.group(2)}"
        from datetime import datetime
        return datetime.now().strftime("%Y-%m")

    def _parse_sheet(
        self,
        ws,
        month: str,
        brand_columns: dict[int, str],
        dispimg_map: dict[tuple[int, int], str],
        img_files: dict[str, str],
        data_start: int,
        category_l1_override: str | None,
        images_dir: str,
    ) -> list[ParsedProduct]:
        """解析 sheet 中的数据行"""
        products = []

        cur_a, cur_b, cur_c, cur_d = "", "", "", ""
        fill_map = self._build_fill_map(ws)
        sorted_brand_cols = sorted(brand_columns.keys())

        for row_idx in range(data_start, ws.max_row + 1):
            cur_a = fill_map.get((row_idx, 1), cur_a)
            cur_b = fill_map.get((row_idx, 2), cur_b)
            cur_c = fill_map.get((row_idx, 3), cur_c)
            cur_d = fill_map.get((row_idx, 4), cur_d)

            if category_l1_override:
                cur_a = category_l1_override

            e_val = ws.cell(row=row_idx, column=5).value
            if not e_val or str(e_val).strip() != "货号":
                continue

            sku_row = row_idx

            # 查找图片行
            image_row = None
            for offset in (1, 2):
                check_row = sku_row - offset
                if check_row >= 1:
                    check_e = ws.cell(row=check_row, column=5).value
                    if check_e and str(check_e).strip() == "图片":
                        image_row = check_row
                        break

            # 查找销量行
            sales_text_map = {}
            sales_row = sku_row + 1
            if sales_row <= ws.max_row:
                next_e = ws.cell(row=sales_row, column=5).value
                if next_e and str(next_e).strip() == "销量":
                    for col_num in sorted_brand_cols:
                        v = ws.cell(row=sales_row, column=col_num).value
                        if v:
                            sales_text_map[col_num] = str(v)

            # 查找升级情况行
            status_text_map = {}
            status_row = sku_row + 2
            if status_row <= ws.max_row:
                next_e2 = ws.cell(row=status_row, column=5).value
                if next_e2 and str(next_e2).strip() == "升级情况":
                    for col_num in sorted_brand_cols:
                        v = ws.cell(row=status_row, column=col_num).value
                        if v:
                            status_text_map[col_num] = str(v)

            # 提取每个品牌列的货号
            for col_num in sorted_brand_cols:
                brand = brand_columns[col_num]
                sku_val = ws.cell(row=sku_row, column=col_num).value
                if not sku_val or not str(sku_val).strip():
                    continue

                # 货号整体保留，换行替换为/
                sku_str = str(sku_val).strip().replace("\r\n", "/").replace("\n", "/")
                sku_str = re.sub(r'/+', '/', sku_str)  # 合并连续斜杠
                sku_str = re.sub(r'\s+', '', sku_str)   # 去除空格

                raw_sales = sales_text_map.get(col_num, "")
                status_text = status_text_map.get(col_num, "")
                status = infer_status_from_text(status_text)

                # 查找对应图片
                image_source = None
                image_path = None
                if image_row and dispimg_map and img_files:
                    # 在该品牌列的图片行查找
                    img_fname = dispimg_map.get((image_row, col_num))

                    if not img_fname:
                        # 对于合并品牌列（如护膝F+G=TMT），图片可能在同品牌的其他列
                        brand_cols = [c for c, b in brand_columns.items() if b == brand]
                        for bc in sorted(brand_cols):
                            if dispimg_map.get((image_row, bc)):
                                img_fname = dispimg_map[(image_row, bc)]
                                break

                    if not img_fname:
                        # 该图片行任意列的图片（兜底）
                        for bc in sorted_brand_cols:
                            if dispimg_map.get((image_row, bc)):
                                img_fname = dispimg_map[(image_row, bc)]
                                break

                    if img_fname and img_fname in img_files:
                        image_source = img_files[img_fname]
                        cat1_clean = sanitize_filename(cur_a)
                        sku_clean = sanitize_filename(sku_str)
                        ext = get_image_ext(img_fname)
                        new_name = f"{cat1_clean}_{sku_clean}{ext}"
                        image_path = f"data/images/{new_name}"

                product = ParsedProduct(
                    category_l1=cur_a,
                    category_l2=cur_b,
                    category_l3=cur_c,
                    version=cur_d if cur_d and cur_d != "暂无" else "",
                    brand=brand,
                    sku=sku_str,
                    sales_volume=parse_sales_volume(raw_sales),
                    status=status,
                    image_path=image_path,
                    raw_sales_text=raw_sales,
                    month=month,
                    _image_source=image_source,
                )
                products.append(product)

        # 重命名图片
        self._rename_images(products, images_dir)
        return products

    def _rename_images(self, products: list[ParsedProduct], images_dir: str):
        """将图片从临时目录复制并重命名为 {品类}_{sku}.ext 到最终目录"""
        import shutil
        import tempfile
        os.makedirs(images_dir, exist_ok=True)
        renamed_set = set()
        for p in products:
            if not p.image_path or not p._image_source:
                continue
            if not os.path.exists(p._image_source):
                continue
            target_name = os.path.basename(p.image_path)
            target_path = os.path.join(images_dir, target_name)
            if target_path not in renamed_set:
                if not os.path.exists(target_path):
                    shutil.copy2(p._image_source, target_path)
                renamed_set.add(target_path)
            p._image_source = None

    def _build_fill_map(self, ws) -> dict:
        """构建合并单元格的向下填充映射（只处理A/B/C/D列）"""
        fill_map = {}
        for mc in ws.merged_cells.ranges:
            if mc.min_col > 4:
                continue
            value = ws.cell(row=mc.min_row, column=mc.min_col).value
            if value is None:
                continue
            for row in range(mc.min_row, mc.max_row + 1):
                fill_map[(row, mc.min_col)] = str(value).strip().replace("\n", "").replace("\r", "")
        return fill_map
