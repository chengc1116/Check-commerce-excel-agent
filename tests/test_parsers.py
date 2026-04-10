# -*- coding: utf-8 -*-
"""
单元测试 — 解析引擎 + 评估引擎 + 文本工具

运行方式:
    pytest tests/ -v
"""

import os
import sys
import tempfile
from pathlib import Path

import pytest

# 确保项目根目录在路径中
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from product_review_agent.utils.text_parser import (
    clean_cell_text,
    extract_segment_id_name,
    parse_demographic,
    parse_duration,
    parse_percentage,
    split_tags,
)
from product_review_agent.parsers.table_classifier import identify_table_type
from product_review_agent.parsers.excel_parser import ExcelParser
from product_review_agent.evaluators.persona_evaluator import PersonaEvaluator
from product_review_agent.evaluators.scenario_evaluator import ScenarioEvaluator
from product_review_agent.evaluators.report_generator import ReportGenerator
from product_review_agent.models import EvaluationDimension, CheckSeverity, CheckItem
from product_review_agent.connectors.market_data import MarketDataConnector, MarketDataInput
from product_review_agent.connectors.search_connector import (
    SearchConnector,
    SearchRequest,
    SearchCategory,
)


# ==============================================================
# 文本解析工具测试
# ==============================================================

class TestParsePercentage:
    """百分比解析测试"""

    def test_standard_percentage(self):
        assert parse_percentage("40%") == 0.4

    def test_decimal_percentage(self):
        assert parse_percentage("12.5%") == 0.125

    def test_no_percent_sign(self):
        assert parse_percentage("40") == 0.4

    def test_small_value(self):
        assert parse_percentage("0.5%") == 0.005

    def test_na_value(self):
        assert parse_percentage("N/A") is None

    def test_empty_value(self):
        assert parse_percentage("") is None
        assert parse_percentage(None) is None

    def test_dash_value(self):
        assert parse_percentage("-") is None


class TestParseDuration:
    """时长解析测试"""

    def test_range_minutes(self):
        result = parse_duration("30-60分钟")
        assert result == {"min": 30, "max": 60, "unit": "分钟"}

    def test_range_hours(self):
        result = parse_duration("1-2小时")
        assert result == {"min": 60, "max": 120, "unit": "分钟"}

    def test_single_minute(self):
        result = parse_duration("30分钟")
        assert result == {"min": 30, "max": 30, "unit": "分钟"}

    def test_na(self):
        assert parse_duration("N/A") is None

    def test_empty(self):
        assert parse_duration("") is None


class TestParseDemographic:
    """人群画像解析测试"""

    def test_age_gender_standard(self):
        result = parse_demographic("30-40岁女性")
        assert result["age_range"] == "30-40"
        assert result["gender"] == "female"

    def test_age_gender_male(self):
        result = parse_demographic("25-35岁男性")
        assert result["age_range"] == "25-35"
        assert result["gender"] == "male"

    def test_age_only(self):
        result = parse_demographic("20-30岁")
        assert result["age_range"] == "20-30"
        assert result["gender"] == "不限"

    def test_text_only(self):
        result = parse_demographic("年轻白领")
        assert result["age_range"] == "未知"
        assert result["gender"] == "不限"


class TestSplitTags:
    """标签拆分测试"""

    def test_slash_separator(self):
        assert split_tags("日常跑步/慢跑") == ["日常跑步", "慢跑"]

    def test_plus_separator(self):
        assert split_tags("减震+稳定") == ["减震", "稳定"]

    def test_comma_separator(self):
        assert split_tags("透气,吸汗,轻便") == ["透气", "吸汗", "轻便"]

    def test_chinese_comma(self):
        assert split_tags("跑步、健身、徒步") == ["跑步", "健身", "徒步"]

    def test_empty(self):
        assert split_tags("") == []
        assert split_tags(None) == []


class TestExtractSegmentIdName:
    """序号名称提取测试"""

    def test_dot_format(self):
        assert extract_segment_id_name("1. 跑步运动") == (1, "跑步运动")

    def test_chinese_dot(self):
        assert extract_segment_id_name("2 健身训练") == (2, "健身训练")

    def test_paren_format(self):
        assert extract_segment_id_name("3、户外徒步") == (3, "户外徒步")

    def test_no_id(self):
        assert extract_segment_id_name("跑步运动") == (None, "跑步运动")

    def test_empty(self):
        assert extract_segment_id_name("") == (None, "")


class TestCleanCellText:
    """文本清洗测试"""

    def test_normal(self):
        assert clean_cell_text("  hello  ") == "hello"

    def test_newlines(self):
        assert clean_cell_text("line1\nline2") == "line1 line2"

    def test_none(self):
        assert clean_cell_text(None) == ""


# ==============================================================
# 表格分类器测试
# ==============================================================

class TestTableClassifier:
    def test_audience_table(self):
        headers = ["5类细分", "占比", "年龄/性别", "核心场景"]
        assert identify_table_type(headers).value == "audience_segmentation"

    def test_scenario_table(self):
        headers = ["具体场景", "占比", "人群", "核心需求", "使用时长"]
        assert identify_table_type(headers).value == "scenario_matrix"

    def test_unknown_table(self):
        headers = ["姓名", "电话", "地址"]
        assert identify_table_type(headers).value == "unknown"

    def test_empty_headers(self):
        assert identify_table_type([]).value == "unknown"


# ==============================================================
# Excel 解析器集成测试
# ==============================================================

class TestExcelParser:
    """Excel 解析器测试（需要示例文件）"""

    @pytest.fixture
    def sample_file(self, tmp_path):
        """生成测试用 Excel 文件"""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "人群分析"

        # 章节标题
        ws.merge_cells("A1:D1")
        ws["A1"] = "4.人群"

        # 表头
        for col, h in enumerate(["5类细分", "占比", "年龄/性别", "核心场景"], 1):
            ws.cell(row=2, column=col, value=h)

        # 数据
        data = [
            ["1. 跑步运动", "40%", "30-40岁女性", "日常跑步/慢跑"],
            ["2. 健身训练", "25%", "25-35岁男性", "健身房力量训练/有氧运动"],
            ["3. 日常通勤", "15%", "20-30岁女性", "上班通勤/日常出行"],
        ]
        for r_idx, row in enumerate(data, 3):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        # 场景表
        ws2 = wb.create_sheet("场景分类")
        ws2.merge_cells("A1:E1")
        ws2["A1"] = "5.场景分类"

        for col, h in enumerate(["具体场景", "占比", "人群", "核心需求", "使用时长"], 1):
            ws2.cell(row=2, column=col, value=h)

        data2 = [
            ["1. 跑步场景", "40%", "跑步运动", "减震+稳定+轻便", "30-60分钟"],
            ["2. 健身场景", "25%", "健身训练", "支撑+防滑+耐磨", "60-90分钟"],
        ]
        for r_idx, row in enumerate(data2, 3):
            for c_idx, val in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=val)

        path = tmp_path / "test.xlsx"
        wb.save(str(path))
        return path

    def test_parse_excel(self, sample_file):
        """测试完整 Excel 解析流程"""
        parser = ExcelParser()
        result = parser.parse(sample_file)

        assert result.document_meta.file_type.value == "excel"
        assert len(result.document_meta.sheets_processed) == 2

        # 验证人群数据
        persona = result.content.get("target_audience")
        assert persona is not None
        assert len(persona["segments"]) == 3

        seg1 = persona["segments"][0]
        assert seg1["name"] == "跑步运动"
        assert seg1["proportion"] == 0.4
        assert seg1["demographic"]["age_range"] == "30-40"
        assert seg1["demographic"]["gender"] == "female"
        assert "日常跑步" in seg1["core_scenarios"]

        # 验证场景数据
        scenarios = result.content.get("usage_scenarios")
        assert scenarios is not None
        assert len(scenarios["scenarios"]) == 2

        sc1 = scenarios["scenarios"][0]
        assert sc1["name"] == "跑步场景"
        assert sc1["proportion"] == 0.4
        assert sc1["target_segment"] == "跑步运动"
        assert "减震" in sc1["core_needs"]
        assert sc1["usage_duration"]["min"] == 30
        assert sc1["usage_duration"]["max"] == 60


# ==============================================================
# 评估引擎测试
# ==============================================================

class TestPersonaEvaluator:
    """人群评估器测试"""

    def setup_method(self):
        self.evaluator = PersonaEvaluator()

    def test_complete_data(self):
        """完整数据应得高分"""
        data = {
            "segments": [
                {
                    "id": 1, "name": "跑步运动", "proportion": 0.4,
                    "demographic": {"age_range": "30-40", "gender": "female"},
                    "core_scenarios": ["日常跑步", "慢跑"],
                },
                {
                    "id": 2, "name": "健身训练", "proportion": 0.6,
                    "demographic": {"age_range": "25-35", "gender": "male"},
                    "core_scenarios": ["健身房训练"],
                },
            ]
        }
        result = self.evaluator.evaluate(data)
        assert result.dimension == "使用人群评估"
        assert result.score > 80

    def test_empty_segments(self):
        """空数据应得低分"""
        data = {"segments": []}
        result = self.evaluator.evaluate(data)
        assert result.score < 50
        has_error = any(c.severity == CheckSeverity.ERROR for c in result.checks)
        assert has_error

    def test_proportion_sum_error(self):
        """占比总和偏离100%应产生ERROR"""
        data = {
            "segments": [
                {"id": 1, "name": "A", "proportion": 0.3,
                 "demographic": {"age_range": "20-30", "gender": "male"}},
                {"id": 2, "name": "B", "proportion": 0.2,
                 "demographic": {"age_range": "30-40", "gender": "female"}},
            ]
        }
        result = self.evaluator.evaluate(data)
        prop_check = next(
            (c for c in result.checks if "占比总和" in c.rule_name), None
        )
        assert prop_check is not None
        assert prop_check.severity == CheckSeverity.ERROR


class TestScenarioEvaluator:
    """场景评估器测试"""

    def setup_method(self):
        self.evaluator = ScenarioEvaluator()

    def test_complete_data(self):
        """完整数据应得高分"""
        data = {
            "scenarios": [
                {
                    "name": "跑步场景", "proportion": 0.4,
                    "target_segment": "跑步运动",
                    "core_needs": ["减震", "稳定"],
                    "usage_duration": {"min": 30, "max": 60, "unit": "分钟"},
                },
                {
                    "name": "健身场景", "proportion": 0.6,
                    "target_segment": "健身训练",
                    "core_needs": ["支撑", "防滑"],
                    "usage_duration": {"min": 60, "max": 90, "unit": "分钟"},
                },
            ]
        }
        result = self.evaluator.evaluate(data)
        assert result.dimension == "使用场景评估"
        assert result.score > 80

    def test_cross_reference_check(self):
        """场景关联不存在的人群应产生WARNING"""
        data = {
            "scenarios": [
                {
                    "name": "跑步场景", "proportion": 0.5,
                    "target_segment": "不存在的分类",
                    "core_needs": ["减震"],
                },
            ]
        }
        audiences = [
            {"name": "跑步运动", "proportion": 0.5},
        ]
        result = self.evaluator.evaluate(data, audiences)
        has_warning = any(
            "关联" in c.rule_name and c.severity == CheckSeverity.WARNING
            for c in result.checks
        )
        assert has_warning


# ==============================================================
# 预留接口测试
# ==============================================================

class TestMarketDataConnector:
    def test_pending_status(self):
        connector = MarketDataConnector()
        result = connector.verify_market_size(MarketDataInput(
            category="跑步鞋",
            claimed_market_size=1500,
            claimed_growth_rate=12.5,
        ))
        assert result.status == "pending"
        assert result.verified_size is None

    def test_validation_error(self):
        connector = MarketDataConnector()
        result = connector.verify_market_size(MarketDataInput(
            category="",
            claimed_market_size=1500,
        ))
        assert result.status == "validation_error"


class TestSearchConnector:
    def test_pending_status(self):
        connector = SearchConnector()
        result = connector.search(SearchRequest(
            query="跑步鞋市场",
            category=SearchCategory.MARKET_SIZE,
        ))
        assert result.status == "pending"
        assert result.total_results == 0

    def test_empty_query(self):
        connector = SearchConnector()
        result = connector.search(SearchRequest(query=""))
        assert result.status == "error"

    def test_convenience_methods(self):
        connector = SearchConnector()
        # 各便捷方法应正常返回
        r1 = connector.search_market_size("跑步鞋")
        assert r1.status == "pending"

        r2 = connector.search_competitors("跑步鞋")
        assert r2.status == "pending"

        r3 = connector.search_scenarios("跑步鞋", ["跑步"])
        assert r3.status == "pending"

        r4 = connector.search_personas("跑步鞋", ["女性"])
        assert r4.status == "pending"


# ==============================================================
# 报告生成器测试
# ==============================================================

class TestReportGenerator:
    def test_generate_report(self):
        from product_review_agent.models import ParsedDocument, DocumentMeta, DocumentFileType

        gen = ReportGenerator()

        parsed = ParsedDocument(
            document_meta=DocumentMeta(
                file_type=DocumentFileType.EXCEL,
                file_name="test.xlsx",
            ),
            content={},
        )

        dim1 = EvaluationDimension(
            dimension="使用人群评估",
            score=85.0,
            checks=[
                CheckItem(rule_id="P001", rule_name="test", severity=CheckSeverity.PASS, message="OK"),
            ],
        )
        dim2 = EvaluationDimension(
            dimension="使用场景评估",
            score=90.0,
            checks=[
                CheckItem(rule_id="S001", rule_name="test", severity=CheckSeverity.PASS, message="OK"),
            ],
        )

        report = gen.generate(parsed, [dim1, dim2])
        assert 80 <= report.overall_score <= 95
        assert "test.xlsx" in report.summary

        text = gen.format_report_text(report)
        assert "Review Report" in text
        assert "使用人群评估" in text
