# -*- coding: utf-8 -*-
"""
解析模块销量Excel → 写入 module_monthly_sales 表

用法:
  python scripts/import_module_sales.py --file "data/excel/模块销量TOP10.xlsx" --month 2026-03
"""

import sys
import io
import os
import argparse
import sqlite3
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = PROJECT_ROOT / "data" / "project_review.db"

CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS module_monthly_sales (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    month           TEXT NOT NULL,
    cbb_code        VARCHAR(20) NOT NULL,
    cbb_name        VARCHAR(100),
    category        VARCHAR(20),
    reuse_area      VARCHAR(50),
    module_sales    INTEGER,
    rank            INTEGER,
    product_code    VARCHAR(20),
    product_name    VARCHAR(100),
    brand           VARCHAR(50),
    product_category VARCHAR(50),
    created_at      DATETIME DEFAULT CURRENT_TIMESTAMP,
    UNIQUE(month, cbb_code, product_code)
);

CREATE INDEX IF NOT EXISTS idx_mms_month ON module_monthly_sales(month);
CREATE INDEX IF NOT EXISTS idx_mms_cbb_code ON module_monthly_sales(cbb_code);
CREATE INDEX IF NOT EXISTS idx_mms_category ON module_monthly_sales(category);
CREATE INDEX IF NOT EXISTS idx_mms_rank ON module_monthly_sales(month, rank);
"""


def parse_excel(file_path: str, month: str) -> list[dict]:
    """解析模块销量Excel"""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # 找到主sheet（第一个sheet或包含"模块"的sheet）
    target_sheet = None
    for name in wb.sheetnames:
        if "模块" in name or "TOP" in name:
            target_sheet = wb[name]
            break
    if not target_sheet:
        target_sheet = wb[wb.sheetnames[0]]

    ws = target_sheet
    print(f"解析Sheet: {ws.title} ({ws.max_row}行)")

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = {}
        for c in row:
            if c.value is not None:
                import re
                m = re.match(r"([A-Z]+)(\d+)", c.coordinate)
                if m:
                    col_letter = m.group(1)
                    vals[col_letter] = c.value

        rank = vals.get("A")
        product_code = vals.get("B")
        if rank is None or product_code is None:
            continue

        module_code = vals.get("L", "")
        if not module_code or "新工厂" in str(module_code):
            continue

        # 从 cbb_modules 查 cbb_name 和 category
        cbb_name = str(vals.get("M", "")).strip()
        category = ""

        records.append({
            "month": month,
            "cbb_code": str(module_code).strip(),
            "cbb_name": cbb_name,
            "category": category,
            "reuse_area": str(vals.get("W", "")).strip() if vals.get("W") else "",
            "module_sales": int(vals.get("AB", 0)) if vals.get("AB") else 0,
            "rank": int(rank),
            "product_code": str(product_code).strip(),
            "product_name": str(vals.get("C", "")).strip() if vals.get("C") else "",
            "brand": str(vals.get("D", "")).strip() if vals.get("D") else "",
            "product_category": str(vals.get("E", "")).strip() if vals.get("E") else "",
        })

    print(f"解析到 {len(records)} 条记录")
    return records


def import_to_db(records: list[dict], db_path: str = None):
    """写入数据库"""
    if db_path is None:
        db_path = str(DB_PATH)

    conn = sqlite3.connect(db_path)
    for stmt in CREATE_TABLE_SQL.split(";"):
        stmt = stmt.strip()
        if stmt:
            conn.execute(stmt)

    # 先补充 cbb_name 和 category（从 cbb_modules 表查询）
    cur = conn.cursor()
    for r in records:
        cur.execute(
            "SELECT cbb_name, category FROM cbb_modules WHERE cbb_code = ?",
            (r["cbb_code"],),
        )
        row = cur.fetchone()
        if row:
            r["cbb_name"] = row[0] or r["cbb_name"]
            r["category"] = row[1] or ""

    # 先删除同月份的数据（覆盖更新）
    if records:
        month = records[0]["month"]
        cur.execute("DELETE FROM module_monthly_sales WHERE month = ?", (month,))
        deleted = cur.rowcount
        if deleted:
            print(f"删除 {month} 旧数据 {deleted} 条")

    # 插入
    insert_sql = """
    INSERT OR REPLACE INTO module_monthly_sales
    (month, cbb_code, cbb_name, category, reuse_area, module_sales, rank,
     product_code, product_name, brand, product_category)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """

    for r in records:
        cur.execute(insert_sql, (
            r["month"], r["cbb_code"], r["cbb_name"], r["category"],
            r["reuse_area"], r["module_sales"], r["rank"],
            r["product_code"], r["product_name"], r["brand"], r["product_category"],
        ))

    conn.commit()
    print(f"成功写入 {len(records)} 条记录")

    # 验证
    cur.execute("SELECT COUNT(*) FROM module_monthly_sales WHERE month = ?", (records[0]["month"],))
    count = cur.fetchone()[0]
    print(f"表中 {records[0]['month']} 数据: {count} 条")

    conn.close()


def main():
    parser = argparse.ArgumentParser(description="导入模块销量Excel到数据库")
    parser.add_argument("--file", required=True, help="模块销量Excel文件路径")
    parser.add_argument("--month", required=True, help="数据月份，如 2026-03")
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"文件不存在: {args.file}")
        return

    records = parse_excel(args.file, args.month)
    if records:
        import_to_db(records)
    else:
        print("无有效数据")


if __name__ == "__main__":
    main()
